# -*- coding: utf-8 -*-
"""
컴활 2급 실기 자동 채점 프로그램 (diff 기반)

사용법:
    python grade.py --problem 문제.xlsm --answer 정답.xlsm --student 내풀이.xlsm \
                    [--key 기대값.json] [--html report.html]

원리:
    문제 파일과 정답 파일의 차이(diff)를 채점 항목으로 자동 추출한 뒤,
    학생 파일이 정답과 일치하는지 검사합니다.

의존성: Python 3.8+ / openpyxl (표준 라이브러리 외 유일한 의존성)
"""

__version__ = "1.4.0"

import argparse
import html as html_mod
import io
import json
import math
import os
import re
import sys
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, time

try:
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
except ImportError:
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(encoding="utf-8")
        except Exception:
            pass
    _MSG = ("오류: 채점에 필요한 openpyxl 라이브러리가 설치되어 있지 않습니다.\n"
            "설치 방법: 명령 프롬프트에서 다음 명령을 실행하세요.\n"
            "    py -m pip install openpyxl\n"
            "  (위 명령이 안 되면: python -m pip install openpyxl)")
    print(_MSG)
    try:
        print(_MSG, file=sys.stderr)
    except Exception:
        pass
    sys.exit(3)  # exit 3 = openpyxl 미설치 (런처가 자동 설치 흐름으로 연결)

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------

DEFAULT_POINTS = {
    "기본작업-1": 5,
    "기본작업-2": 10,
    "기본작업-3": 5,
    "계산작업": 40,
    "분석작업-1": 10,
    "분석작업-2": 10,
    "매크로작업": 10,
    "차트작업": 10,
}
PASS_LINE = 70
CALC_GROUP_TARGET = 5      # 계산작업 문제 수
MAX_SCAN_ROWS = 400        # diff 스캔 상한 (안전장치)
MAX_SCAN_COLS = 100

NS_C = "http://schemas.openxmlformats.org/drawingml/2006/chart"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS = {"c": NS_C, "a": NS_A}

FMT_KIND_LABEL = {
    "value": "셀 값",
    "number_format": "표시 형식",
    "font": "글꼴 서식",
    "fill": "채우기",
    "alignment": "맞춤(정렬)",
    "border": "테두리",
    "merge": "셀 병합",
    "rowheight": "행 높이",
    "colwidth": "열 너비",
    "names": "정의된 이름",
    "style": "셀 스타일",
}

# ---------------------------------------------------------------------------
# 공통 유틸
# ---------------------------------------------------------------------------


def norm_sheet_name(name):
    """시트명 정규화: 공백 제거, '정답'/'-정답' 접미사 제거."""
    n = re.sub(r"\s+", "", str(name))
    for suf in ("-정답", "(정답)", "_정답", "정답"):
        if n.endswith(suf) and len(n) > len(suf):
            n = n[: -len(suf)]
            break
    return n


def formula_text(raw):
    """셀 raw 값에서 수식 문자열을 추출. 수식이 아니면 None."""
    if isinstance(raw, str) and raw.startswith("="):
        return raw
    if isinstance(raw, ArrayFormula):
        return raw.text if raw.text else "=__ARRAY__"
    if isinstance(raw, DataTableFormula):
        return "=__DATATABLE__"
    return None


def norm_formula(f):
    """수식 정규화. 문자열 리터럴은 보존, 나머지는 대문자화·공백/$/' 제거.
    RANK( 는 RANK.EQ( 와 동일 취급, _xlfn. 접두사 제거, 시트명 속 '정답' 제거."""
    if f is None:
        return None
    if not isinstance(f, str):
        f = str(f)
    f = f.lstrip("=")
    parts = re.split(r'("(?:[^"]|"")*")', f)
    out = []
    for i, p in enumerate(parts):
        if i % 2 == 1:  # 문자열 리터럴 그대로
            out.append(p)
        else:
            p = p.upper()
            p = re.sub(r"\s+", "", p)
            p = p.replace("$", "").replace("'", "")
            p = p.replace("_XLFN.", "")
            p = p.replace("정답!", "!")  # '계산작업 정답'! -> 계산작업!
            p = re.sub(r"\[\d+\]", "", p)  # 외부 참조 [1]시트명! 제거
            p = p.replace("RANK.EQ(", "RANK(")
            out.append(p)
    return "".join(out)


def _as_number(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _as_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        m = re.match(
            r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})"
            r"(?:[ T](\d{1,2}):(\d{2})(?::(\d{2}))?)?$", v.strip())
        if m:
            try:
                return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                                int(m.group(4) or 0), int(m.group(5) or 0),
                                int(m.group(6) or 0))
            except ValueError:
                return None
    return None


def value_eq(a, b):
    """관대한 값 비교. 숫자: rel_tol=1e-6 + round(x,6), 문자열: strip,
    날짜/시간: 타입 변환, None과 빈 문자열은 동일."""
    a_empty = a is None or (isinstance(a, str) and not a.strip())
    b_empty = b is None or (isinstance(b, str) and not b.strip())
    if a_empty or b_empty:
        return a_empty and b_empty
    if isinstance(a, bool) or isinstance(b, bool):
        return a is b or a == b
    na, nb = _as_number(a), _as_number(b)
    if na is not None and nb is not None:
        if math.isclose(na, nb, rel_tol=1e-6, abs_tol=1e-9):
            return True
        return round(na, 6) == round(nb, 6)
    da, db = _as_dt(a), _as_dt(b)
    if da is not None and db is not None:
        return da == db
    if isinstance(a, time) and isinstance(b, time):
        return a == b
    if isinstance(a, str) and isinstance(b, str):
        return a.strip() == b.strip()
    return a == b


def fmt_value(v):
    """리포트 표시용 값 문자열."""
    if v is None:
        return "(비어 있음)"
    if isinstance(v, float):
        r = round(v, 6)
        if r == int(r):
            return str(int(r))
        return str(r)
    if isinstance(v, datetime):
        if v.hour == v.minute == v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, (ArrayFormula, DataTableFormula)):
        return "(배열/데이터표 수식)"
    s = str(v)
    if len(s) > 60:
        s = s[:57] + "..."
    return s


def disp_width(s):
    w = 0
    for ch in s:
        w += 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
    return w


def pad(s, width, align="left"):
    gap = width - disp_width(s)
    if gap <= 0:
        return s
    if align == "right":
        return " " * gap + s
    if align == "center":
        l = gap // 2
        return " " * l + s + " " * (gap - l)
    return s + " " * gap


# ---------------------------------------------------------------------------
# 워크북 래퍼
# ---------------------------------------------------------------------------


class Book:
    """수식용/캐시값용 워크북과 zip 패키지 정보를 함께 보관."""

    def __init__(self, path, label):
        self.path = path
        self.label = label
        if not os.path.isfile(path):
            raise FileNotFoundError(f"{label} 파일을 찾을 수 없습니다: {path}")
        try:
            self.raw = openpyxl.load_workbook(path, data_only=False)
            self.cached = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            raise RuntimeError(f"{label} 파일을 열 수 없습니다 ({path}): {e}")
        self.zf = zipfile.ZipFile(path)
        self.sheet_part = self._map_sheet_parts()
        self.norm_map = {}
        for s in self.raw.sheetnames:
            self.norm_map.setdefault(norm_sheet_name(s), s)

    def _map_sheet_parts(self):
        """시트명 -> xl/worksheets/sheetN.xml 매핑."""
        mapping = {}
        try:
            wbxml = self.zf.read("xl/workbook.xml").decode("utf-8", "replace")
            relsxml = self.zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
            rels = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', relsxml))
            for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml):
                name = html_mod.unescape(m.group(1))
                target = rels.get(m.group(2), "")
                if target.startswith("/"):
                    part = target.lstrip("/")
                else:
                    part = "xl/" + target
                part = part.replace("xl/xl/", "xl/")
                mapping[name] = part
        except Exception:
            pass
        return mapping

    def sheet_xml(self, sheet_name):
        part = self.sheet_part.get(sheet_name)
        if not part:
            return None
        try:
            return self.zf.read(part).decode("utf-8", "replace")
        except KeyError:
            return None

    def sheet_rel_targets(self, sheet_name):
        """시트의 관계(rels) 대상 파트 경로 목록."""
        part = self.sheet_part.get(sheet_name)
        if not part:
            return []
        rels_part = re.sub(r"worksheets/(sheet\d+\.xml)$", r"worksheets/_rels/\1.rels", part)
        try:
            x = self.zf.read(rels_part).decode("utf-8", "replace")
        except KeyError:
            return []
        out = []
        for m in re.finditer(r'Target="([^"]+)"', x):
            t = m.group(1)
            if t.startswith("/"):
                out.append(t.lstrip("/"))
            else:
                out.append(os.path.normpath("xl/worksheets/" + t).replace("\\", "/"))
        return out

    def has_vba(self):
        return any("vbaProject" in n for n in self.zf.namelist())

    def has_pivot_for_sheet(self, sheet_name):
        return any("pivotTable" in t for t in self.sheet_rel_targets(sheet_name))

    def has_scenarios(self, sheet_name):
        x = self.sheet_xml(sheet_name)
        return bool(x) and "<scenarios" in x

    def sheet_has_drawing(self, sheet_name):
        """단추/도형 존재 판정: drawing, vmlDrawing, 컨트롤 참조."""
        for t in self.sheet_rel_targets(sheet_name):
            if "/drawings/" in t or "vmlDrawing" in t or "ctrlProp" in t:
                return True
        x = self.sheet_xml(sheet_name) or ""
        return "<controls" in x or "<legacyDrawing" in x

    def chart_parts_for_sheet(self, sheet_name):
        """시트 -> drawing -> chart 파트 경로 목록 (rels 체인)."""
        charts = []
        for t in self.sheet_rel_targets(sheet_name):
            m = re.match(r"xl/drawings/(drawing\d+)\.xml$", t)
            if not m:
                continue
            drels = f"xl/drawings/_rels/{m.group(1)}.xml.rels"
            try:
                x = self.zf.read(drels).decode("utf-8", "replace")
            except KeyError:
                continue
            for mm in re.finditer(r'Target="([^"]+charts/chart\d+\.xml)"', x):
                p = mm.group(1)
                p = p.lstrip("/") if p.startswith("/") else \
                    os.path.normpath("xl/drawings/" + p).replace("\\", "/")
                charts.append(p)
        if not charts:  # 폴백: 워크북 내 아무 차트
            charts = sorted(
                n for n in self.zf.namelist() if re.match(r"xl/charts/chart\d+\.xml$", n)
            )
        return charts


# ---------------------------------------------------------------------------
# 서식 시그니처
# ---------------------------------------------------------------------------


# 색 비교: 테마/인덱스 색을 RGB로 환산한 뒤 비교 (저장 방식 차이 무시)

_THEME_ORDER = ("lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3",
                "accent4", "accent5", "accent6", "hlink", "folHlink")
_DEFAULT_THEME = ("FFFFFF", "000000", "E7E6E6", "44546A", "4472C4", "ED7D31",
                  "A5A5A5", "FFC000", "5B9BD5", "70AD47", "0563C1", "954F72")
_theme_cache = {}


def _theme_palette(wb):
    """워크북 theme1.xml -> 테마 색 RGB 목록 (Excel theme 인덱스 순)."""
    key = id(wb)
    if key in _theme_cache:
        return _theme_cache[key]
    palette = list(_DEFAULT_THEME)
    try:
        xml = getattr(wb, "loaded_theme", None)
        if xml:
            if isinstance(xml, bytes):
                xml = xml.decode("utf-8", "replace")
            found = {}
            for name in ("dk1", "lt1", "dk2", "lt2", "accent1", "accent2",
                         "accent3", "accent4", "accent5", "accent6",
                         "hlink", "folHlink"):
                m = re.search(
                    rf"<a:{name}>.*?(?:srgbClr\s+val=\"([0-9A-Fa-f]{{6}})\""
                    rf"|sysClr[^>]*lastClr=\"([0-9A-Fa-f]{{6}})\")",
                    xml, re.DOTALL)
                if m:
                    found[name] = (m.group(1) or m.group(2)).upper()
            for i, nm in enumerate(_THEME_ORDER):
                if nm in found:
                    palette[i] = found[nm]
    except Exception:
        pass
    _theme_cache[key] = palette
    return palette


def _apply_tint(rgb_hex, tint):
    """테마 색 tint 적용 (표준 HLS 휘도 공식)."""
    import colorsys
    try:
        r = int(rgb_hex[0:2], 16) / 255.0
        g = int(rgb_hex[2:4], 16) / 255.0
        b = int(rgb_hex[4:6], 16) / 255.0
        h, l, s = colorsys.rgb_to_hls(r, g, b)
        if tint < 0:
            l = l * (1.0 + tint)
        elif tint > 0:
            l = l * (1.0 - tint) + tint
        r2, g2, b2 = colorsys.hls_to_rgb(h, max(0.0, min(1.0, l)), s)
        return "%02X%02X%02X" % (round(r2 * 255), round(g2 * 255),
                                 round(b2 * 255))
    except Exception:
        return rgb_hex


def _quant_hex(rgb_hex):
    """반올림 오차 흡수를 위해 채널을 4단위로 양자화."""
    try:
        vals = [min(255, (int(rgb_hex[i:i + 2], 16) + 2) // 4 * 4)
                for i in (0, 2, 4)]
        return "%02X%02X%02X" % tuple(vals)
    except Exception:
        return rgb_hex


def _color_key(color, wb=None):
    """색을 저장 방식과 무관한 비교 키로. 가능하면 ("rgb", 양자화 RGB)."""
    if color is None:
        return None
    try:
        if color.type == "rgb":
            argb = color.rgb
            if argb in (None, "00000000"):
                return None
            return ("rgb", _quant_hex(str(argb)[-6:].upper()))
        if color.type == "theme":
            idx = color.theme
            tint = float(color.tint or 0.0)
            if wb is not None and isinstance(idx, int) and \
                    0 <= idx < len(_THEME_ORDER):
                base = _theme_palette(wb)[idx]
                return ("rgb", _quant_hex(_apply_tint(base, tint)))
            return ("theme", idx, round(tint, 2))
        if color.type == "indexed":
            if color.indexed in (64, 65):  # 자동/기본
                return None
            try:
                from openpyxl.styles.colors import COLOR_INDEX
                argb = COLOR_INDEX[color.indexed]
                return ("rgb", _quant_hex(str(argb)[-6:].upper()))
            except Exception:
                return ("idx", color.indexed)
    except Exception:
        pass
    return None


_COLOR_NAMES_KO = (
    ("FFFFFF", "흰색"), ("000000", "검정"), ("FF0000", "빨강"),
    ("FFFF00", "노랑"), ("FFC000", "황금색"), ("BF8F00", "황금색(어둡게)"),
    ("ED7D31", "주황"), ("00B050", "초록"), ("70AD47", "연한 초록(강조6)"),
    ("A9D08E", "연한 초록"), ("4472C4", "파랑(강조1)"), ("0070C0", "파랑"),
    ("5B9BD5", "하늘색(강조5)"), ("A5A5A5", "회색(강조3)"),
    ("D9D9D9", "연한 회색"), ("808080", "회색"), ("7030A0", "보라"),
    ("FFE699", "연한 황금색"), ("DDEBF7", "연한 파랑"),
    ("FCE4D6", "연한 주황"),
)


def color_name_ko(rgb_hex):
    """RGB -> 가까운 한국어 색 이름 (+hex). 예: '황금색(#BF8F00 계열)'"""
    try:
        r = int(rgb_hex[0:2], 16)
        g = int(rgb_hex[2:4], 16)
        b = int(rgb_hex[4:6], 16)
        best = None
        for hx, nm in _COLOR_NAMES_KO:
            d = (abs(r - int(hx[0:2], 16)) + abs(g - int(hx[2:4], 16))
                 + abs(b - int(hx[4:6], 16)))
            if best is None or d < best[0]:
                best = (d, nm)
        if best and best[0] <= 90:
            return f"{best[1]}(#{rgb_hex} 계열)"
        return f"#{rgb_hex}"
    except Exception:
        return f"#{rgb_hex}"


def fmt_signature(cell, kind):
    """셀의 특정 서식 종류(kind) 시그니처. 색은 테마 해석 후 RGB 비교."""
    try:
        wb = None
        try:
            wb = cell.parent.parent
        except Exception:
            pass
        if kind == "number_format":
            nf = (cell.number_format or "General").strip()
            return nf
        if kind == "font":
            f = cell.font
            u = f.underline
            u = None if u in (None, "none") else u
            return (f.name, float(f.size or 11), bool(f.bold), bool(f.italic),
                    u, bool(f.strike), _color_key(f.color, wb))
        if kind == "fill":
            fl = cell.fill
            pat = fl.patternType if fl else None
            if pat in (None, "none"):
                return None
            return (pat, _color_key(fl.fgColor, wb))
        if kind == "alignment":
            al = cell.alignment
            return (al.horizontal, al.vertical, bool(al.wrap_text),
                    int(al.text_rotation or 0))
        if kind == "border":
            b = cell.border
            return tuple(bool(s is not None and s.style) for s in
                         (b.left, b.right, b.top, b.bottom, b.diagonal))
    except Exception:
        return "?"
    return None


# 셀 단위 비교 서식 종류. 오탐 방지를 위해 다음은 제외:
#  - border: edge(격자 선) 단위 합성 비교로 대체 (인접 셀 공유 선 대응)
#  - 셀 스타일 이름: 저장 방식 차이일 뿐, 효과는 표시형식/글꼴 등으로 판정됨
CELL_FMT_KINDS = ("number_format", "font", "fill", "alignment")

_BORDER_WEIGHT = {"hair": 1, "thin": 2, "dotted": 2, "dashed": 2,
                  "dashDot": 2, "dashDotDot": 2, "mediumDashed": 3,
                  "mediumDashDot": 3, "mediumDashDotDot": 3,
                  "slantDashDot": 3, "medium": 3, "thick": 4, "double": 5}

BORDER_STYLE_KO = {"thin": "실선", "medium": "굵은 실선", "double": "이중 실선",
                   "dotted": "점선", "dashed": "파선", "hair": "가는 실선",
                   "thick": "매우 굵은 실선", "mediumDashed": "굵은 파선",
                   "dashDot": "일점쇄선", "dashDotDot": "이점쇄선"}


def _combine_edge(s1, s2):
    """인접한 두 셀이 선언한 같은 선: 어느 쪽이든 있으면 존재, 굵은 쪽 우선."""
    if s1 and s2:
        return s1 if _BORDER_WEIGHT.get(s1, 2) >= _BORDER_WEIGHT.get(s2, 2) \
            else s2
    return s1 or s2


def sheet_edge_map(ws, max_r, max_c):
    """격자 선(edge) 단위 테두리 맵. Excel이 공유 선을 어느 셀에 저장했든
    같은 선으로 취급. 병합 범위 내부 선은 제외.

    키: ("h", r, c) = r행 위쪽 가로선, ("v", r, c) = c열 왼쪽 세로선.
    """
    _members, anchors = merge_maps(ws)

    def side(r, c, name):
        if r < 1 or c < 1:
            return None
        try:
            s = getattr(ws.cell(r, c).border, name)
            return s.style if s is not None else None
        except Exception:
            return None

    edges = {}
    for r in range(1, max_r + 2):
        for c in range(1, max_c + 1):
            v = _combine_edge(side(r, c, "top"), side(r - 1, c, "bottom"))
            if v:
                edges[("h", r, c)] = v
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 2):
            v = _combine_edge(side(r, c, "left"), side(r, c - 1, "right"))
            if v:
                edges[("v", r, c)] = v
    for (min_c, min_r, mx_c, mx_r) in anchors.values():
        for r in range(min_r + 1, mx_r + 1):
            for c in range(min_c, mx_c + 1):
                edges.pop(("h", r, c), None)
        for r in range(min_r, mx_r + 1):
            for c in range(min_c + 1, mx_c + 1):
                edges.pop(("v", r, c), None)
    return edges


def merge_maps(ws):
    """병합 맵: (비앵커 멤버 -> 앵커 좌표), (앵커 -> bounds).

    Excel은 서식 적용 순서(병합 전/후)에 따라 병합 하위 셀에 서식을
    저장하기도, 앵커에만 저장하기도 합니다. 화면 표시는 항상 앵커
    기준이므로 서식 비교도 앵커 기준으로 해야 오탐이 없습니다.
    """
    members = {}
    anchors = {}
    try:
        for rng in ws.merged_cells.ranges:
            min_c, min_r, max_c, max_r = rng.bounds
            anchors[(min_r, min_c)] = (min_c, min_r, max_c, max_r)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if (r, c) != (min_r, min_c):
                        members[(r, c)] = (min_r, min_c)
    except Exception:
        pass
    return members, anchors


def fmt_signature_m(ws, r, c, kind, members, anchors):
    """병합 인지 서식 시그니처.

    비앵커 멤버는 앵커 셀의 서식으로 대체하고, 병합 앵커의 테두리는
    범위 외곽 테두리를 합성해 비교합니다 (외곽 테두리가 하위 셀에
    분산 저장되는 경우 대응).
    """
    if (r, c) in members:
        r, c = members[(r, c)]
    if kind == "border" and (r, c) in anchors:
        min_c, min_r, max_c, max_r = anchors[(r, c)]

        def _has(side, cells):
            for rr, cc in cells:
                try:
                    s = getattr(ws.cell(rr, cc).border, side)
                    if s is not None and s.style:
                        return True
                except Exception:
                    pass
            return False

        return (
            _has("left", [(rr, min_c) for rr in range(min_r, max_r + 1)]),
            _has("right", [(rr, max_c) for rr in range(min_r, max_r + 1)]),
            _has("top", [(min_r, cc) for cc in range(min_c, max_c + 1)]),
            _has("bottom", [(max_r, cc) for cc in range(min_c, max_c + 1)]),
            False,
        )
    return fmt_signature(ws.cell(r, c), kind)


# ---------------------------------------------------------------------------
# 서식 속성 한국어 표기 / 조작 경로 힌트 / 범위 표기
# ---------------------------------------------------------------------------

H_ALIGN_KO = {"center": "가운데", "centerContinuous": "선택 영역의 가운데로",
              "distributed": "균등 분할", "fill": "채우기", "left": "왼쪽",
              "right": "오른쪽", "justify": "양쪽 맞춤", "general": "일반",
              None: "일반"}
V_ALIGN_KO = {"center": "가운데", "top": "위쪽", "bottom": "아래쪽",
              "justify": "양쪽", "distributed": "균등 분할", None: "기본"}
UNDERLINE_KO = {"double": "이중 실선 밑줄", "single": "밑줄",
                "doubleAccounting": "이중 실선 밑줄(회계형)",
                "singleAccounting": "밑줄(회계형)"}


def describe_font(sig):
    """글꼴 시그니처 -> '굴림체 16pt 굵은 기울임꼴, 이중 실선 밑줄'"""
    if not sig or sig == "?":
        return "확인 불가"
    name, size, bold, italic, u, strike, color = sig
    head = f"{name or '기본 글꼴'} {size:g}pt"
    parts = []
    if bold and italic:
        parts.append("굵은 기울임꼴")
    elif bold:
        parts.append("굵게")
    elif italic:
        parts.append("기울임꼴")
    if u:
        parts.append(UNDERLINE_KO.get(u, "밑줄"))
    if strike:
        parts.append("취소선")
    if color:
        if isinstance(color, tuple) and color[0] == "rgb":
            parts.append(f"글꼴 색 {color_name_ko(color[1])}")
        else:
            parts.append("글꼴 색 지정")
    return head + (" " + ", ".join(parts) if parts else "")


def describe_border(sig):
    if not sig or sig == "?":
        return "확인 불가"
    l, r, t, b = sig[:4]
    if all((l, r, t, b)):
        return "테두리(사방)"
    if not any((l, r, t, b)):
        return "테두리 없음"
    sides = [nm for nm, v in (("왼쪽", l), ("오른쪽", r),
                              ("위쪽", t), ("아래쪽", b)) if v]
    return "테두리(" + "·".join(sides) + ")"


def describe_fill(sig):
    if sig is None:
        return "채우기 없음"
    if sig == "?":
        return "확인 불가"
    pat, color = sig
    if isinstance(color, tuple) and color and color[0] == "rgb":
        return f"채우기 {color_name_ko(str(color[1])[-6:])}"
    if isinstance(color, tuple) and color and color[0] == "theme":
        return "채우기 있음(테마 색)"
    return "채우기 있음"


def props_for_kind(kind, sig_a, sig_s):
    """실제로 다른 속성만 (한국어 속성명, 정답, 내 답) 목록으로."""
    out = []
    if kind == "alignment":
        sa = sig_a if isinstance(sig_a, tuple) else (None, None, False, 0)
        ss = sig_s if isinstance(sig_s, tuple) else (None, None, False, 0)
        if sa[0] != ss[0]:
            out.append({"name": "가로 맞춤",
                        "expected": H_ALIGN_KO.get(sa[0], sa[0]),
                        "got": H_ALIGN_KO.get(ss[0], ss[0])})
        if sa[1] != ss[1]:
            out.append({"name": "세로 맞춤",
                        "expected": V_ALIGN_KO.get(sa[1], sa[1]),
                        "got": V_ALIGN_KO.get(ss[1], ss[1])})
        if sa[2] != ss[2]:
            out.append({"name": "텍스트 줄 바꿈",
                        "expected": "적용" if sa[2] else "해제",
                        "got": "적용" if ss[2] else "해제"})
        if len(sa) > 3 and len(ss) > 3 and sa[3] != ss[3]:
            out.append({"name": "텍스트 회전",
                        "expected": f"{sa[3]}도", "got": f"{ss[3]}도"})
    elif kind == "font":
        out.append({"name": "글꼴", "expected": describe_font(sig_a),
                    "got": describe_font(sig_s)})
    elif kind == "fill":
        out.append({"name": "채우기", "expected": describe_fill(sig_a),
                    "got": describe_fill(sig_s)})
    elif kind == "border":
        out.append({"name": "테두리", "expected": describe_border(sig_a),
                    "got": describe_border(sig_s)})
    elif kind == "number_format":
        exp = str(sig_a)
        mean = explain_number_format(exp)
        out.append({"name": "표시 형식",
                    "expected": exp + (f" — {mean}" if mean else ""),
                    "got": str(sig_s) if sig_s is not None else "확인 불가"})
    elif kind == "style":
        out.append({"name": "셀 스타일", "expected": str(sig_a),
                    "got": str(sig_s)})
    return out


def hints_for_kind(kind, sig_a):
    """실제 차이 난 속성에서 파생한 조작 경로 힌트."""
    if kind == "alignment" and isinstance(sig_a, tuple):
        h = sig_a[0]
        if h == "distributed":
            return ["Ctrl+1 → 맞춤 탭 → 가로 '균등 분할'"]
        if h == "centerContinuous":
            return ["Ctrl+1 → 맞춤 탭 → 가로 '선택 영역의 가운데로' "
                    "(병합이 아닙니다)"]
        if h == "fill":
            return ["Ctrl+1 → 맞춤 탭 → 가로 '채우기'"]
        if sig_a[2]:
            return ["Ctrl+1 → 맞춤 탭 → '텍스트 줄 바꿈' 체크"]
        return ["홈 탭 맞춤 그룹(가로/세로 가운데)"]
    if kind == "font" and isinstance(sig_a, tuple):
        if sig_a[4] in ("double", "doubleAccounting"):
            return ["Ctrl+1 → 글꼴 탭 → 밑줄 '이중 실선' (회계용 아님 주의)",
                    "홈 탭 글꼴 그룹에서 글꼴 이름·크기·굵게/기울임 지정"]
        return ["홈 탭 글꼴 그룹에서 글꼴 이름·크기·굵게/기울임 지정"]
    if kind == "merge":
        return ["범위를 선택하고 홈 탭 → '병합하고 가운데 맞춤'"]
    if kind == "number_format":
        return [f"Ctrl+1 → 표시 형식 → 사용자 지정: {sig_a}"]
    if kind == "border":
        return ["홈 탭 → 테두리 → 모든 테두리/바깥쪽 테두리"]
    if kind == "fill":
        return ["홈 탭 → 채우기 색(페인트 통)"]
    if kind == "style":
        return ["홈 탭 → 셀 스타일에서 지정(쉼표/통화 등)"]
    if kind == "value":
        return ["문제 지시 값을 정확히 입력했는지 확인"]
    return []


def compress_coords(coords):
    """['C5','C6',...,'C11','H6'] -> ['C5:C11', 'H6'] 범위 표기."""
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_to_tuple
    pts = set()
    for co in coords:
        try:
            if ":" in str(co):
                pts.add(str(co))  # 이미 범위 표기
                continue
            pts.add(coordinate_to_tuple(str(co)))
        except Exception:
            pts.add(str(co))
    ranges = sorted(p for p in pts if isinstance(p, str))
    cells = sorted(p for p in pts if isinstance(p, tuple))
    out = []
    used = set()
    bycol = {}
    for (r, c) in cells:
        bycol.setdefault(c, []).append(r)
    for c in sorted(bycol):
        rows = sorted(bycol[c])
        run = [rows[0]]
        for r in rows[1:] + [None]:
            if r is not None and r == run[-1] + 1:
                run.append(r)
                continue
            if len(run) >= 2:
                L = get_column_letter(c)
                out.append(f"{L}{run[0]}:{L}{run[-1]}")
                used.update((rr, c) for rr in run)
            if r is not None:
                run = [r]
    remaining = [p for p in cells if p not in used]
    byrow = {}
    for (r, c) in remaining:
        byrow.setdefault(r, []).append(c)
    for r in sorted(byrow):
        cols = sorted(byrow[r])
        run = [cols[0]]
        for c in cols[1:] + [None]:
            if c is not None and c == run[-1] + 1:
                run.append(c)
                continue
            if len(run) >= 2:
                out.append(f"{get_column_letter(run[0])}{r}:"
                           f"{get_column_letter(run[-1])}{r}")
                used.update((r, cc) for cc in run)
            elif (r, run[0]) not in used:
                out.append(f"{get_column_letter(run[0])}{r}")
                used.add((r, run[0]))
            if c is not None:
                run = [c]
    return sorted(set(out + ranges))


def _cluster_units(units, target):
    """서식 diff 단위를 공간 인접(맨해튼 거리 <=2)으로 묶고 target개로 병합."""
    if not units:
        return []
    clusters = [[u] for u in units]

    def dist(c1, c2):
        return min(abs(a["pos"][0] - b["pos"][0])
                   + abs(a["pos"][1] - b["pos"][1])
                   for a in c1 for b in c2)

    merged = True
    while merged:
        merged = False
        for i in range(len(clusters)):
            for j in range(i + 1, len(clusters)):
                if dist(clusters[i], clusters[j]) <= 2:
                    clusters[i].extend(clusters[j])
                    del clusters[j]
                    merged = True
                    break
            if merged:
                break
    while len(clusters) > max(1, target):
        best = None
        for i in range(len(clusters)):
            for j in range(i + 1, len(clusters)):
                d = dist(clusters[i], clusters[j])
                if best is None or d < best[0]:
                    best = (d, i, j)
        _d, i, j = best
        clusters[i].extend(clusters[j])
        del clusters[j]
    for cl in clusters:
        cl.sort(key=lambda u: (u["pos"][0], u["pos"][1]))
    clusters.sort(key=lambda cl: (cl[0]["pos"][0], cl[0]["pos"][1]))
    return clusters


def norm_ref(ref):
    """참조 문자열 정규화: $, 공백, 따옴표, '정답' 제거, 대문자화."""
    if ref is None:
        return None
    s = str(ref).upper().replace("$", "").replace("'", "")
    s = re.sub(r"\s+", "", s)
    s = s.replace("정답!", "!").replace("정답".upper() + "!", "!")
    return s


def norm_range_only(ref):
    """참조에서 시트명을 떼고 범위만 (예: '차트작업 정답'!$C$4:$E$4 -> C4:E4)."""
    if ref is None:
        return None
    s = str(ref)
    if "!" in s:
        s = s.rsplit("!", 1)[1]
    return s.upper().replace("$", "").replace("'", "").strip()


# ---------------------------------------------------------------------------
# 조건부 서식
# ---------------------------------------------------------------------------


def cf_rule_set(ws):
    """시트의 조건부 서식 규칙을 정규화된 튜플 집합으로."""
    rules = set()
    try:
        for cf in ws.conditional_formatting:
            sqref = norm_ref(str(cf.sqref))
            for r in cf.rules:
                formulas = ()
                try:
                    if getattr(r, "formula", None):
                        formulas = tuple(sorted(norm_formula("=" + f) for f in r.formula))
                except Exception:
                    pass
                rules.add((sqref, r.type, getattr(r, "operator", None), formulas))
    except Exception:
        pass
    return rules


# ---------------------------------------------------------------------------
# diff 추출
# ---------------------------------------------------------------------------


def scan_bounds(*sheets):
    mr = max([s.max_row or 1 for s in sheets] + [1])
    mc = max([s.max_column or 1 for s in sheets] + [1])
    return min(mr, MAX_SCAN_ROWS), min(mc, MAX_SCAN_COLS)


def cell_differs(raw_p, raw_a):
    """문제/정답 셀 raw 값이 채점 대상이 될 만큼 다른가."""
    fp, fa = formula_text(raw_p), formula_text(raw_a)
    if fp is not None or fa is not None:
        if (fp is None) != (fa is None):
            return True
        return norm_formula(fp) != norm_formula(fa)
    return not value_eq(raw_p, raw_a)


def value_diff_cells(shp, sha):
    """문제↔정답 값/수식 diff 셀 좌표 목록 [(row, col, coord)]."""
    mr, mc = scan_bounds(shp, sha)
    out = []
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            if cell_differs(shp.cell(r, c).value, sha.cell(r, c).value):
                out.append((r, c, sha.cell(r, c).coordinate))
    return out


def format_diff_items(book_p, book_a, psheet, asheet):
    """문제↔정답 서식 diff를 종류별 항목으로. {kind: [셀좌표...] 또는 데이터}"""
    shp, sha = book_p.raw[psheet], book_a.raw[asheet]
    mr, mc = scan_bounds(shp, sha)
    items = {}
    p_members, p_anchors = merge_maps(shp)
    a_members, a_anchors = merge_maps(sha)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            # 병합 하위(비앵커) 셀은 앵커가 대표하므로 개별 비교하지 않음
            if (r, c) in a_members or (r, c) in p_members:
                continue
            for kind in CELL_FMT_KINDS:
                sp = fmt_signature_m(shp, r, c, kind, p_members, p_anchors)
                sa = fmt_signature_m(sha, r, c, kind, a_members, a_anchors)
                if sp != sa:
                    items.setdefault(kind, []).append(
                        sha.cell(r, c).coordinate)
    # 테두리: edge(격자 선) 단위 diff — 인접 셀에 분산 저장돼도 같은 선
    from openpyxl.utils import get_column_letter
    p_edges = sheet_edge_map(shp, mr, mc)
    a_edges = sheet_edge_map(sha, mr, mc)
    border_edges = {}
    for key in set(p_edges) | set(a_edges):
        av, pv = a_edges.get(key), p_edges.get(key)
        if av != pv:
            ek, r, c = key
            rr, cc = min(r, mr), min(c, mc)
            coord = f"{get_column_letter(cc)}{rr}"
            border_edges.setdefault(coord, []).append((key, av))
    if border_edges:
        items["border"] = sorted(border_edges)
        items["border_edges"] = border_edges
    # 병합
    merges_p = {str(x).upper() for x in shp.merged_cells.ranges}
    merges_a = {str(x).upper() for x in sha.merged_cells.ranges}
    if merges_p != merges_a:
        items["merge"] = sorted(merges_a - merges_p) or sorted(merges_a ^ merges_p)
    # 행 높이 / 열 너비 (±0.5)
    rows = set(shp.row_dimensions) | set(sha.row_dimensions)
    rh = []
    for r in sorted(rows):
        hp = shp.row_dimensions[r].height if r in shp.row_dimensions else None
        ha = sha.row_dimensions[r].height if r in sha.row_dimensions else None
        if _size_differs(hp, ha):
            rh.append((r, ha))
    if rh:
        items["rowheight"] = rh
    cols = set(shp.column_dimensions) | set(sha.column_dimensions)
    cw = []
    for cl in sorted(cols):
        wp = shp.column_dimensions[cl].width if cl in shp.column_dimensions else None
        wa = sha.column_dimensions[cl].width if cl in sha.column_dimensions else None
        if _size_differs(wp, wa):
            cw.append((cl, wa))
    if cw:
        items["colwidth"] = cw
    # 정의된 이름 (이 시트를 가리키는 것)
    names = defined_names_for_sheet(book_a, asheet)
    names_p = defined_names_for_sheet(book_p, psheet)
    new_names = {k: v for k, v in names.items() if names_p.get(k) != v}
    if new_names:
        items["names"] = new_names
    return items


def _size_differs(a, b):
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    return abs(a - b) > 0.5


def _size_matches(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= 0.5


def defined_names_for_sheet(book, sheet_name):
    """이 시트를 참조 대상에 포함하는 정의된 이름 {이름: 정규화참조}."""
    out = {}
    key = norm_sheet_name(sheet_name)
    try:
        for name, dn in book.raw.defined_names.items():
            val = norm_ref(dn.value)
            if val and key.upper() in norm_sheet_name(val.split("!")[0]).upper():
                out[str(name)] = val
    except Exception:
        pass
    return out


# ---------------------------------------------------------------------------
# 셀 채점
# ---------------------------------------------------------------------------


class CellJudge:
    """기대값/기대수식 대비 학생 셀 판정.

    key_cells: {좌표: {"value": 기대값 or None, "formula": 기대수식 or None}}
    """

    def __init__(self, book_a, book_s, asheet, ssheet, key_cells):
        self.sha_raw = book_a.raw[asheet]
        self.sha_val = book_a.cached[asheet]
        self.shs_raw = book_s.raw[ssheet] if ssheet else None
        self.shs_val = book_s.cached[ssheet] if ssheet else None
        self.key_cells = key_cells or {}
        self.formula_only_used = False  # 캐시값 없어 수식 판정한 적 있음

    def expected(self, r, c):
        """(기대값 or __MISSING__, 기대수식) 반환."""
        coord = self.sha_raw.cell(r, c).coordinate
        raw = self.sha_raw.cell(r, c).value
        f = formula_text(raw)
        ov = self.key_cells.get(coord)
        if ov is not None:
            if ov.get("formula"):
                f = ov["formula"]
            if ov.get("value") is not None:
                return ov["value"], f
        if f is not None:
            cached = self.sha_val.cell(r, c).value
            if cached is not None:
                return cached, f
            return "__MISSING__", f
        return raw, None

    def student(self, r, c):
        """(학생값 or __MISSING__, 학생수식)"""
        if self.shs_raw is None:
            return "__MISSING__", None
        raw = self.shs_raw.cell(r, c).value
        f = formula_text(raw)
        if f is not None:
            cached = self.shs_val.cell(r, c).value
            if cached is not None:
                return cached, f
            return "__MISSING__", f
        return raw, None

    def judge(self, r, c):
        """(정답 여부, 기대값, 학생값, 기대수식) 반환."""
        exp, ef = self.expected(r, c)
        got, sf = self.student(r, c)
        if self.shs_raw is None:
            return False, exp, None, ef
        # 1) 값 판정
        if exp != "__MISSING__" and got != "__MISSING__":
            ok = value_eq(exp, got)
            if not ok and ef and sf and norm_formula(ef) == norm_formula(sf):
                # 수식이 정확히 같은데 캐시값만 다른 경우(재계산 안 됨) 관대 처리
                ok = True
            return ok, exp, got, ef
        # 2) 값 판정 불가 -> 수식 판정
        if ef is not None or sf is not None:
            self.formula_only_used = True
            ok = ef is not None and sf is not None and \
                norm_formula(ef) == norm_formula(sf)
            return ok, exp if exp != "__MISSING__" else None, sf or got, ef
        # 3) 둘 다 수식 없음
        e = None if exp == "__MISSING__" else exp
        g = None if got == "__MISSING__" else got
        return value_eq(e, g), e, g, ef


# ---------------------------------------------------------------------------
# --key JSON 정규화
# ---------------------------------------------------------------------------


def expand_ref(ref):
    """'D3' 또는 'D3:D12' 참조를 (row, col) 목록으로 확장."""
    from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
    ref = str(ref).replace("$", "").replace(" ", "").upper()
    if ":" in ref:
        mn_c, mn_r, mx_c, mx_r = range_boundaries(ref)
        return [(r, c) for r in range(mn_r, mx_r + 1)
                for c in range(mn_c, mx_c + 1)]
    r, c = coordinate_to_tuple(ref)
    return [(r, c)]


def _unwrap_cell_override(v):
    """key cells 항목을 {"value":..., "formula":...} 형태로 통일."""
    if isinstance(v, dict) and ("value" in v or "formula" in v):
        return {"value": v.get("value"), "formula": v.get("formula")}
    return {"value": v, "formula": None}


def normalize_key(key):
    """--key JSON을 내부 표준 형태로 정규화.

    cells는 두 가지 형태를 모두 지원:
      중첩형: {"시트명": {"A1": 값 또는 {"value":..,"formula":..}}}
      평면형: {"시트명!A1": 값 또는 {"value":..,"formula":..}}
    groups의 항목은 단일 좌표("D3")와 범위("D3:D12") 모두 허용.
    """
    out = {"points": {}, "cells": {}, "groups": {}}
    for k, v in (key.get("points") or {}).items():
        out["points"][norm_sheet_name(k)] = v
    for k, v in (key.get("cells") or {}).items():
        if "!" in str(k):  # 평면형
            sheet, coord = str(k).rsplit("!", 1)
            out["cells"].setdefault(norm_sheet_name(sheet), {})[
                coord.replace("$", "").strip().upper()] = _unwrap_cell_override(v)
        elif isinstance(v, dict):  # 중첩형
            tgt = out["cells"].setdefault(norm_sheet_name(k), {})
            for coord, cv in v.items():
                tgt[str(coord).replace("$", "").strip().upper()] = \
                    _unwrap_cell_override(cv)
    for k, v in (key.get("groups") or {}).items():
        out["groups"][norm_sheet_name(k)] = v
    out["format_checks"] = {}
    for k, v in (key.get("format_checks") or {}).items():
        if isinstance(v, list):
            out["format_checks"][norm_sheet_name(k)] = v
    return out


# ---------------------------------------------------------------------------
# 계산작업 그룹핑
# ---------------------------------------------------------------------------


def cluster_cells(cells, target):
    """diff 셀을 인접(체비쇼프 거리 1) 기준으로 묶고, target 개수로 병합."""
    if not cells:
        return []
    clusters = [[cell] for cell in cells]

    def dist(c1, c2):
        return min(
            max(abs(a[0] - b[0]), abs(a[1] - b[1])) for a in c1 for b in c2
        )

    # 1단계: 인접 셀 병합
    merged = True
    while merged:
        merged = False
        for i in range(len(clusters)):
            for j in range(i + 1, len(clusters)):
                if dist(clusters[i], clusters[j]) <= 1:
                    clusters[i].extend(clusters[j])
                    del clusters[j]
                    merged = True
                    break
            if merged:
                break
    # 2단계: target 개수 초과 시 가장 가까운 클러스터 병합
    while len(clusters) > target:
        best = None
        for i in range(len(clusters)):
            for j in range(i + 1, len(clusters)):
                d = dist(clusters[i], clusters[j])
                if best is None or d < best[0]:
                    best = (d, i, j)
        _, i, j = best
        clusters[i].extend(clusters[j])
        del clusters[j]
    # 좌표 정렬
    for cl in clusters:
        cl.sort(key=lambda x: (x[0], x[1]))
    clusters.sort(key=lambda cl: (cl[0][0], cl[0][1]))
    return clusters


# ---------------------------------------------------------------------------
# 차트 특성 추출
# ---------------------------------------------------------------------------


def chart_features(book, sheet_name):
    """시트에 연결된 첫 차트의 특성 dict. 차트가 없으면 None."""
    parts = book.chart_parts_for_sheet(sheet_name)
    if not parts:
        return None
    try:
        root = ET.fromstring(book.zf.read(parts[0]))
    except Exception:
        return None
    plot = root.find(".//c:plotArea", NS)
    if plot is None:
        return None
    types = []
    series = []
    for child in plot:
        tag = child.tag.split("}")[1]
        if not tag.endswith("Chart"):
            continue
        types.append(tag)
        for ser in child.findall("c:ser", NS):
            ref = None
            for path in (".//c:val/c:numRef/c:f", ".//c:yVal/c:numRef/c:f",
                         ".//c:val//c:f"):
                e = ser.find(path, NS)
                if e is not None:
                    ref = e.text
                    break
            has_dlbls = ser.find("c:dLbls", NS) is not None
            series.append((norm_range_only(ref), has_dlbls))
    # 제목
    title = None
    t = root.find("c:chart/c:title", NS)
    autodel = root.find("c:chart/c:autoTitleDeleted", NS)
    if t is not None and (autodel is None or autodel.get("val") in ("0", "false", None)):
        texts = [e.text or "" for e in t.findall(".//a:t", NS)]
        txt = "".join(texts).strip()
        if not txt:
            f = t.find(".//c:strRef/c:f", NS)
            txt = norm_range_only(f.text) if f is not None else ""
        title = ("있음", txt)
    legend = root.find("c:chart/c:legend/c:legendPos", NS)
    legend_pos = legend.get("val") if legend is not None else (
        "r" if root.find("c:chart/c:legend", NS) is not None else None)
    axes = []
    for ax in root.findall(".//c:valAx", NS):
        mx = ax.find("c:scaling/c:max", NS)
        mn = ax.find("c:scaling/c:min", NS)
        mju = ax.find("c:majorUnit", NS)
        axes.append((
            float(mx.get("val")) if mx is not None else None,
            float(mn.get("val")) if mn is not None else None,
            float(mju.get("val")) if mju is not None else None,
        ))
    return {
        "types": sorted(types),
        "title": title,
        "series": series,
        "legend": legend_pos,
        "axes": axes,
    }


# ---------------------------------------------------------------------------
# 시트별 채점
# ---------------------------------------------------------------------------


class SheetResult:
    def __init__(self, name, alloc):
        self.name = name
        self.alloc = alloc
        self.earned = 0
        self.details = []   # 틀린 항목 설명 목록 (콘솔용 텍스트)
        self.notes = []     # 안내/경고
        self.missing = False
        self.wrong = []     # 오답노트 카드(구조화) 목록


def make_cell_entries(judge, wrong, limit=3):
    """오답 셀들의 (좌표, 기대값, 학생값, 기대/학생 수식) 구조화 목록."""
    out = []
    for (r, c, coord) in wrong[:limit]:
        ok, exp, got, ef = judge.judge(r, c)
        try:
            sf = judge.student(r, c)[1]
        except Exception:
            sf = None
        out.append({"coord": coord, "expected": fmt_value(exp),
                    "got": fmt_value(got), "formula": ef,
                    "got_formula": sf})
    return out


def add_card(res, label, lost, kind, cells=None, formula=None,
             note=None, hint=None, more=0, props=None):
    """오답노트 카드 1장 추가."""
    res.wrong.append({
        "sheet": res.name, "label": label, "lost": round(float(lost), 1),
        "kind": kind, "cells": cells or [], "formula": formula,
        "note": note, "hint": hint, "more": more, "props": props or [],
        "category": None, "explain": [], "point": None,
    })


def add_wrong_cells(res, judge, wrong, limit=30):
    for (r, c, coord) in wrong[:limit]:
        ok, exp, got, ef = judge.judge(r, c)
        line = f"{coord}: 기대값 {fmt_value(exp)} / 학생값 {fmt_value(got)}"
        if ef:
            line += f" / 기대 수식 {ef}"
        res.details.append(line)
    if len(wrong) > limit:
        res.details.append(f"... 외 {len(wrong) - limit}개 셀")


def grade_basic1(res, ctx):
    """기본작업-1: 전부 일치=만점, 하나라도 틀리면 0점."""
    diffs = ctx["vdiffs"]
    if not diffs:
        res.earned = res.alloc
        res.notes.append("문제/정답 간 차이가 없어 자동 만점 처리")
        return
    judge = ctx["judge"]
    wrong = [(r, c, co) for (r, c, co) in diffs if not judge.judge(r, c)[0]]
    if wrong:
        res.earned = 0
        res.details.append(f"자료 입력 {len(diffs)}개 셀 중 {len(wrong)}개 불일치 "
                           "(부분점수 없음, 0점)")
        add_wrong_cells(res, judge, wrong)
        add_card(res, "자료 입력", res.alloc, "cell",
                 cells=make_cell_entries(judge, wrong, 5),
                 more=max(0, len(wrong) - 5),
                 hint="기본작업-1은 부분점수가 없습니다. 입력을 마친 뒤 "
                      "문제지와 셀 단위로 1:1 대조하는 습관을 들이세요.")
    else:
        res.earned = res.alloc


def _edge_desc(key, coord):
    """edge 키 -> 'H11 아래쪽' 같은 위치 설명."""
    from openpyxl.utils.cell import coordinate_to_tuple
    try:
        ek, r, c = key
        tr, tc = coordinate_to_tuple(coord)
        if ek == "h":
            return f"{coord} {'아래쪽' if r > tr else '위쪽'}"
        return f"{coord} {'오른쪽' if c > tc else '왼쪽'}"
    except Exception:
        return str(coord)


def _border_props(fail_units, s_edges):
    """테두리 실패 유닛들 -> (정답 설명, 내 답 설명)."""
    styles = set()
    misses = []
    for u in fail_units:
        for key, style in u.get("edges") or []:
            if style:
                styles.add(style)
            sv = s_edges.get(key) if s_edges is not None else None
            if (sv or None) != (style or None):
                if style and not sv:
                    misses.append(f"{_edge_desc(key, u['coord'])} 선 없음")
                elif style and sv:
                    misses.append(f"{_edge_desc(key, u['coord'])} 선 종류 다름"
                                  f"({BORDER_STYLE_KO.get(sv, sv)})")
                else:
                    misses.append(f"{_edge_desc(key, u['coord'])} 불필요한 선")
    if len(styles) == 1:
        exp = f"모든 지시 선({BORDER_STYLE_KO.get(next(iter(styles)), '실선')})"
    elif styles:
        exp = "테두리 적용(" + "·".join(
            BORDER_STYLE_KO.get(s, s) for s in sorted(styles)) + ")"
    else:
        exp = "선 제거"
    got = ", ".join(misses[:3]) + (f" 외 {len(misses) - 3}곳"
                                   if len(misses) > 3 else "")
    return exp, (got or "확인 불가")


def _key_row_heights(ctx, sheet_name):
    """--key format_checks에 명시된 행 높이 {행번호: pt}."""
    out = {}
    fcs = (ctx["key"].get("format_checks") or {}).get(
        norm_sheet_name(sheet_name)) or []
    for ent in fcs:
        chk = (ent or {}).get("check") or {}
        rh = chk.get("row_height", chk.get("row_heights"))
        if rh is None:
            continue
        if isinstance(rh, dict):
            for k, v in rh.items():
                try:
                    out[int(k)] = float(v)
                except (TypeError, ValueError):
                    pass
            continue
        rows = []
        rng = str(ent.get("range") or "").strip()
        m = re.match(r"^(\d+)(?:\s*[:~-]\s*(\d+))?\s*행?$", rng)
        if m:
            rows = range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1)
        else:
            try:
                from openpyxl.utils.cell import range_boundaries
                _c1, r1, _c2, r2 = range_boundaries(rng)
                rows = range(r1, r2 + 1)
            except Exception:
                pass
        for r in rows:
            try:
                out[int(r)] = float(rh)
            except (TypeError, ValueError):
                pass
    return out


def _key_col_widths(ctx, sheet_name):
    """--key format_checks에 명시된 열 너비 {열문자: 너비}."""
    out = {}
    fcs = (ctx["key"].get("format_checks") or {}).get(
        norm_sheet_name(sheet_name)) or []
    for ent in fcs:
        chk = (ent or {}).get("check") or {}
        cw = chk.get("col_width", chk.get("col_widths"))
        if cw is None:
            continue
        if isinstance(cw, dict):
            for k, v in cw.items():
                try:
                    out[str(k).strip().upper().rstrip("열")] = float(v)
                except (TypeError, ValueError):
                    pass
            continue
        cols = []
        rng = str(ent.get("range") or "").strip().upper()
        m = re.match(r"^([A-Z]{1,3})(?:\s*[:~-]\s*([A-Z]{1,3}))?\s*열?$", rng)
        if m:
            from openpyxl.utils import column_index_from_string, \
                get_column_letter
            c1 = column_index_from_string(m.group(1))
            c2 = column_index_from_string(m.group(2) or m.group(1))
            cols = [get_column_letter(i) for i in range(c1, c2 + 1)]
        for cl in cols:
            try:
                out[cl] = float(cw)
            except (TypeError, ValueError):
                pass
    return out


def grade_basic2(res, ctx):
    """기본작업-2: 서식 diff를 '지시 단위'(공간 인접 클러스터)로 채점.

    - 채점 후보는 문제↔정답 diff 셀만 (지시 범위 밖 잔여 서식 감점 없음)
    - 병합은 앵커 기준, 테두리는 edge 합성, 색은 테마 해석 후 RGB 비교
    - 행 높이/열 너비는 자동 조정 부산물이 많아 기본 제외(참고 노트만),
      --key format_checks에 명시된 세트만 채점(±0.5)
    """
    from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
    book_a, book_s = ctx["book_a"], ctx["book_s"]
    asheet, ssheet = ctx["asheet"], ctx["ssheet"]
    judge = ctx["judge"]
    fd = ctx["fdiffs"]
    sha = book_a.raw[asheet]
    shs = book_s.raw[ssheet] if ssheet else None
    a_members, a_anchors = merge_maps(sha)
    s_members, s_anchors = merge_maps(shs) if shs is not None else ({}, {})

    # --- 채점 단위(unit) 구성: 문제↔정답 diff에서만 ---
    units = []
    for kind in CELL_FMT_KINDS:
        for coord in fd.get(kind, []):
            r, c = coordinate_to_tuple(coord)
            units.append({"pos": (r, c), "kind": kind, "coord": coord})
    for coord in fd.get("border", []):
        r, c = coordinate_to_tuple(coord)
        units.append({"pos": (r, c), "kind": "border", "coord": coord,
                      "edges": (fd.get("border_edges") or {}).get(coord, [])})
    for rng in fd.get("merge", []):
        try:
            min_c, min_r, _mc, _mr = range_boundaries(rng)
        except Exception:
            continue
        units.append({"pos": (min_r, min_c), "kind": "merge",
                      "range": rng, "coord": rng})
    for (r, c, coord) in ctx["vdiffs"]:
        units.append({"pos": (r, c), "kind": "value", "coord": coord,
                      "rc": (r, c)})

    # --- 별도 항목: 정의된 이름 / (key 명시 시) 행 높이 ---
    standalone = []
    if fd.get("names"):
        standalone.append(("names", fd["names"]))
    key_rh = _key_row_heights(ctx, res.name)
    if key_rh:
        standalone.append(("key_rowheight", key_rh))
    key_cw = _key_col_widths(ctx, res.name)
    if key_cw:
        standalone.append(("key_colwidth", key_cw))

    # --- 행 높이/열 너비: 감점 없는 참고 노트만 (1pt 초과 차이일 때) ---
    if not key_rh and fd.get("rowheight") and shs is not None:
        noted = []
        for (r, ha) in fd["rowheight"]:
            hs = shs.row_dimensions[r].height \
                if r in shs.row_dimensions else None
            if ha is not None and (hs is None or abs(hs - ha) > 1.0):
                noted.append(str(r))
        if noted:
            res.notes.append(
                f"행 높이가 정답 파일과 다릅니다({', '.join(noted[:6])}행) — "
                "글꼴 크기 등에 따른 자동 조정 차이일 수 있어 감점하지 "
                "않습니다. 문제지에 행 높이 지시가 있는 경우만 확인하세요.")
    if not key_cw and fd.get("colwidth") and shs is not None:
        noted = []
        for (cl, wa) in fd["colwidth"]:
            ws_ = shs.column_dimensions[cl].width \
                if cl in shs.column_dimensions else None
            if wa is not None and (ws_ is None or abs(ws_ - wa) > 1.0):
                noted.append(str(cl))
        if noted:
            res.notes.append(
                f"열 너비가 정답 파일과 다릅니다({', '.join(noted[:6])}열) — "
                "감점하지 않습니다. 문제지에 열 너비 지시가 있는 경우만 "
                "확인하세요.")

    if not units and not standalone:
        res.earned = res.alloc
        res.notes.append("채점할 서식 차이가 없어 만점 처리")
        return

    target = max(1, 5 - len(standalone))
    clusters = _cluster_units(units, target) if units else []
    n = len(clusters) + len(standalone)
    per = res.alloc / n
    passed = 0
    mr, mc = scan_bounds(sha, shs) if shs is not None else scan_bounds(sha)
    s_edges = sheet_edge_map(shs, mr, mc) if shs is not None else {}
    s_merges = {str(x).upper() for x in shs.merged_cells.ranges} \
        if shs is not None else set()

    for cl in clusters:
        fails = []
        for u in cl:
            k = u["kind"]
            if k == "merge":
                if u["range"].upper() not in s_merges:
                    fails.append(u)
            elif k == "value":
                r, c = u["rc"]
                if not judge.judge(r, c)[0]:
                    fails.append(u)
            elif k == "border":
                bad = any((s_edges.get(key) or None) != (style or None)
                          for key, style in (u.get("edges") or []))
                if shs is None or bad:
                    fails.append(u)
            else:
                r, c = u["pos"]
                sig_a = fmt_signature_m(sha, r, c, k, a_members, a_anchors)
                sig_s = fmt_signature_m(shs, r, c, k, s_members, s_anchors) \
                    if shs is not None else None
                if sig_a != sig_s:
                    u["sig_a"], u["sig_s"] = sig_a, sig_s
                    fails.append(u)
        if not fails:
            passed += 1
            continue
        # --- 실패 카드: 범위 표기 + 속성별 정답/내답 + 파생 힌트 ---
        loc = compress_coords([u["coord"] for u in cl])
        props = []
        hints = []
        seen_kinds = []
        for u in fails:
            k = u["kind"]
            if k in seen_kinds and k != "value":
                continue
            seen_kinds.append(k)
            if k == "merge":
                rngs = [x["range"] for x in fails if x["kind"] == "merge"]
                props.append({"name": "병합",
                              "expected": "병합하고 가운데 맞춤 ("
                                          + ", ".join(rngs[:3]) + ")",
                              "got": "병합 없음"})
                hints.extend(hints_for_kind("merge", None))
            elif k == "value":
                if sum(1 for p in props if p["name"].startswith("셀 값")) < 2:
                    r, c = u["rc"]
                    _ok, exp, got, _ef = judge.judge(r, c)
                    props.append({"name": f"셀 값({u['coord']})",
                                  "expected": fmt_value(exp),
                                  "got": fmt_value(got)})
                    hints.extend(hints_for_kind("value", None))
            elif k == "border":
                b_fails = [x for x in fails if x["kind"] == "border"]
                exp, got = _border_props(b_fails, s_edges)
                props.append({"name": "테두리", "expected": exp, "got": got})
                hints.extend(hints_for_kind("border", None))
            else:
                props.extend(props_for_kind(k, u.get("sig_a"), u.get("sig_s")))
                hints.extend(hints_for_kind(k, u.get("sig_a")))
        uniq_hints = []
        for h in hints:
            if h not in uniq_hints:
                uniq_hints.append(h)
        label = "서식 지시 (" + ", ".join(loc[:2]) \
            + (" 외" if len(loc) > 2 else "") + ")"
        add_card(res, label, per, "format",
                 cells=[{"coord": t} for t in loc[:6]],
                 props=props, hint=" / ".join(uniq_hints[:3]) or None,
                 more=max(0, len(loc) - 6))
        res.details.append(
            "[서식] " + ", ".join(loc[:4]) + " 불일치: "
            + ", ".join(dict.fromkeys(p["name"] for p in props)))

    # --- 별도 항목 채점 ---
    for kind, payload in standalone:
        if kind == "names":
            stu_names = defined_names_for_sheet(book_s, ssheet) \
                if ssheet else {}
            missing = [k for k, v in payload.items()
                       if stu_names.get(k) != v]
            if not missing:
                passed += 1
            else:
                res.details.append("[정의된 이름] 불일치 (이름 "
                                   + ", ".join(missing[:5]) + ")")
                add_card(res, "서식 - 정의된 이름", per, "format",
                         cells=[{"coord": nm} for nm in missing[:5]],
                         props=[{"name": "정의된 이름",
                                 "expected": f"'{missing[0]}' = "
                                             + str(payload[missing[0]]),
                                 "got": "없음 또는 다른 범위"}],
                         hint="범위를 선택한 뒤 이름 상자(수식 입력줄 왼쪽)에 "
                              "이름을 입력하고 Enter를 누릅니다.")
        elif kind == "key_rowheight":
            bad = []
            for r, want in sorted(payload.items()):
                hs = shs.row_dimensions[r].height \
                    if shs is not None and r in shs.row_dimensions else None
                if not _size_matches(hs, want):
                    bad.append((r, want, hs))
            if not bad:
                passed += 1
            else:
                rows_txt = ", ".join(str(r) for r, _w, _h in bad) + "행"
                res.details.append(f"[행 높이] 불일치 ({rows_txt})")
                add_card(res, "서식 - 행 높이", per, "format",
                         cells=[{"coord": f"{r}행"} for r, _w, _h in bad],
                         props=[{"name": f"{r}행 높이",
                                 "expected": f"{w:g}",
                                 "got": f"{h:g}" if h is not None else "기본"}
                                for r, w, h in bad[:3]],
                         hint="행 머리글 우클릭 → 행 높이에서 숫자를 "
                              "입력합니다.")
        elif kind == "key_colwidth":
            bad = []
            for cl, want in sorted(payload.items()):
                cw_s = shs.column_dimensions[cl].width \
                    if shs is not None and cl in shs.column_dimensions \
                    else None
                if not _size_matches(cw_s, want):
                    bad.append((cl, want, cw_s))
            if not bad:
                passed += 1
            else:
                cols_txt = ", ".join(cl for cl, _w, _s in bad) + "열"
                res.details.append(f"[열 너비] 불일치 ({cols_txt})")
                add_card(res, "서식 - 열 너비", per, "format",
                         cells=[{"coord": f"{cl}열"} for cl, _w, _s in bad],
                         props=[{"name": f"{cl}열 너비",
                                 "expected": f"{w:g}",
                                 "got": f"{s:g}" if s is not None else "기본"}
                                for cl, w, s in bad[:3]],
                         hint="열 머리글 우클릭 → 열 너비에서 숫자를 "
                              "입력합니다.")

    res.earned = int(round(per * passed))
    if passed < n:
        res.details.insert(0, f"서식 지시 {n}개 중 {passed}개 통과 "
                              f"(항목당 {per:.1f}점)")


def _is_cond_syntax(v):
    return isinstance(v, str) and bool(
        re.match(r"^\s*(<>|<=|>=|<|>|=)", v) or "*" in v or "?" in v)


def _norm_field(v):
    return re.sub(r"\s+", "", str(v)).lower() if v is not None else None


def _norm_cond(v):
    """조건/데이터 값 정규화 (공백·대소문자 무시, 수식은 수식 정규화)."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s.startswith("="):
            return ("f", norm_formula(s))
        return ("s", re.sub(r"\s+", "", s).lower())
    if isinstance(v, bool):
        return ("b", v)
    if isinstance(v, (int, float)):
        return ("n", round(float(v), 6))
    return ("s", str(v))


def _block_grid(sh, r1, r2, c1, c2):
    if sh is None:
        return []
    return [[sh.cell(r, c).value for c in range(c1, c2 + 1)]
            for r in range(r1, r2 + 1)]


def _cond_rows(grid):
    """조건 범위 -> (행별 (필드,조건) 짝 집합의 Counter, 머리글)."""
    from collections import Counter
    header = [_norm_field(v) for v in grid[0]]
    rows = []
    for row in grid[1:]:
        pairs = frozenset(
            (header[i], _norm_cond(v)) for i, v in enumerate(row)
            if v is not None and header[i])
        if pairs:
            rows.append(pairs)
    return Counter(rows), header


def compare_condition_block(ga, gs):
    """고급 필터 조건 범위 의미 비교 — 열 순서/행 순서 무관.

    같은 행의 (필드, 조건) 짝 집합이 행 단위로 일치하면 동일한 조건.
    """
    ca, _ha = _cond_rows(ga)
    cs, _hs = _cond_rows(gs) if gs else (None, None)
    if cs is not None and ca == cs:
        return True, []
    msgs = []
    flat_s = {}
    if cs:
        for rowset in cs:
            for (f, cond) in rowset:
                flat_s.setdefault(f, []).append(cond)

    def _fmt(cond):
        if cond is None:
            return "없음"
        return str(cond[1])

    missing_rows = ca - cs if cs is not None else ca
    for rowset in missing_rows:
        for (f, cond) in sorted(rowset, key=str):
            got = flat_s.get(f)
            if got and cond not in got:
                msgs.append(f"'{f}' 조건이 {_fmt(cond)} 이어야 하는데 "
                            f"{_fmt(got[0])} 입니다")
            elif not got:
                msgs.append(f"'{f}' 조건({_fmt(cond)})이 없습니다 — 필드명 "
                            "오타 여부도 확인하세요")
    if not msgs:
        msgs.append("조건의 AND(같은 행)/OR(다른 행) 배치가 정답과 다릅니다")
    return False, msgs[:4]


def compare_result_block(ga, gs):
    """고급 필터 결과 표 비교 — 열 순서는 머리글로 재정렬, 행은 집합 비교.

    반환: (통과, 메시지들, 행 순서만 다른지)
    """
    from collections import Counter
    ha = [_norm_field(v) for v in ga[0]]
    hs = [_norm_field(v) for v in gs[0]] if gs else []
    ha_set = sorted(h for h in ha if h)
    hs_set = sorted(h for h in hs if h)
    if ha_set != hs_set:
        return False, ["추출된 필드(머리글)가 정답과 다릅니다: 정답 "
                       + ", ".join(ha_set) + " / 내 답 "
                       + (", ".join(hs_set) or "없음")], False
    idx_s = {}
    for i, h in enumerate(hs):
        if h and h not in idx_s:
            idx_s[h] = i
    cols_a = [i for i, h in enumerate(ha) if h]

    def _rows(grid, mapper):
        out = []
        for row in grid[1:]:
            t = tuple(mapper(row, i) for i in cols_a)
            if any(x is not None for x in t):
                out.append(t)
        return out

    rows_a = _rows(ga, lambda row, i: _norm_cond(row[i]))
    rows_s = _rows(gs, lambda row, i: _norm_cond(row[idx_s[ha[i]]]))
    if Counter(rows_a) == Counter(rows_s):
        return True, [], rows_a != rows_s
    only_a = Counter(rows_a) - Counter(rows_s)
    only_s = Counter(rows_s) - Counter(rows_a)
    msgs = []
    if only_a:
        msgs.append(f"정답에 있는 결과 행 {sum(only_a.values())}개가 "
                    "누락되었습니다")
    if only_s:
        msgs.append(f"정답에 없는 행 {sum(only_s.values())}개가 "
                    "포함되었습니다 (조건 재확인)")
    return False, msgs or ["결과 행이 정답과 다릅니다"], False


def grade_basic3(res, ctx):
    """기본작업-3: 조건부 서식 규칙 비교, 없으면 값 diff.

    고급 필터는 의미 동치 비교: 조건 범위의 열/행 배치가 달라도 (필드,
    조건) 짝이 같으면 정답, 결과 표는 행 집합 비교(순서 무관·무감점 노트).
    텍스트 나누기 등 일반 블록은 셀 단위 비교 유지.
    """
    book_p, book_a, book_s = ctx["book_p"], ctx["book_a"], ctx["book_s"]
    psheet, asheet, ssheet = ctx["psheet"], ctx["asheet"], ctx["ssheet"]
    cf_a = cf_rule_set(book_a.raw[asheet])
    cf_p = cf_rule_set(book_p.raw[psheet])
    target = cf_a - cf_p
    if target:
        cf_s = cf_rule_set(book_s.raw[ssheet]) if ssheet else set()
        matched = 0
        for rule in target:
            if rule in cf_s:
                matched += 1
            elif any(r[1:] == rule[1:] for r in cf_s):  # 범위만 다르고 규칙 동일
                matched += 1
                res.notes.append("조건부 서식 적용 범위가 정답과 다르지만 규칙이 "
                                 "같아 정답 처리")
            else:
                sq, typ, op, fs = rule
                res.details.append(
                    f"[조건부 서식] 누락/불일치: 범위 {sq}, 유형 {typ}"
                    + (f", 수식 {'; '.join(fs)}" if fs else ""))
                add_card(res, "조건부 서식", res.alloc / len(target), "cf",
                         formula=("=" + fs[0]) if fs else None,
                         note=f"적용 범위 {sq} / 규칙 유형 {typ}"
                              + (f" / 연산자 {op}" if op else ""),
                         hint="홈 탭 → 조건부 서식 → 새 규칙 → '수식을 사용하여 "
                              "서식을 지정할 셀 결정'. 포인트: 행마다 검사하려면 "
                              "열만 고정하는 혼합참조($E5)를 쓰고, 텍스트 비교는 "
                              "큰따옴표(\"...\")로 감쌉니다.")
        res.earned = int(round(res.alloc * matched / len(target)))
        return
    diffs = ctx["vdiffs"]
    if not diffs:
        res.earned = res.alloc
        res.notes.append("문제/정답 간 차이가 없어 자동 만점 처리")
        return
    judge = ctx["judge"]
    from openpyxl.utils import get_column_letter
    shp = book_p.raw[psheet]
    sha = book_a.raw[asheet]
    shs = book_s.raw[ssheet] if ssheet else None
    coord_of = {(r, c): co for (r, c, co) in diffs}
    clusters = cluster_cells([(r, c) for (r, c, _co) in diffs], target=99)
    # 원본 표 머리글 후보: 문제 시트의 문자열 값들
    header_pool = set()
    try:
        for row in shp.iter_rows(min_row=1,
                                 max_row=min(40, shp.max_row or 1)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    header_pool.add(_norm_field(cell.value))
    except Exception:
        pass
    blocks = []
    for cl in clusters:
        r1 = min(r for r, _c in cl)
        r2 = max(r for r, _c in cl)
        c1 = min(c for _r, c in cl)
        c2 = max(c for _r, c in cl)
        ga = _block_grid(sha, r1, r2, c1, c2)
        first = [v for v in ga[0] if v is not None] if ga else []
        headerish = bool(first) and all(
            isinstance(v, str) and _norm_field(v) in header_pool
            for v in first)
        cond_syntax = any(_is_cond_syntax(v) for row in ga[1:] for v in row)
        if headerish and len(ga) >= 2 and (cond_syntax or len(ga) == 2):
            btype = "cond"
        elif headerish and len(ga) >= 3:
            btype = "result"
        else:
            btype = "cells"
        blocks.append((btype, cl, (r1, r2, c1, c2), ga))
    passed_b = 0
    for btype, cl, (r1, r2, c1, c2), ga in blocks:
        rng_txt = (f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
                   if (r1, c1) != (r2, c2) else f"{get_column_letter(c1)}{r1}")
        gs = _block_grid(shs, r1, r2, c1, c2) if shs is not None else []
        if btype == "cond":
            ok, msgs = compare_condition_block(ga, gs)
            if ok:
                passed_b += 1
                continue
            res.details.append(f"[고급 필터 조건] {rng_txt}: "
                               + "; ".join(msgs))
            add_card(res, f"고급 필터 조건 ({rng_txt})",
                     res.alloc / len(blocks), "cell",
                     cells=[{"coord": rng_txt}],
                     props=[{"name": "조건", "expected": "문제 지시 조건",
                             "got": m} for m in msgs[:3]],
                     hint="필드명은 원본 표 머리글과 똑같이, AND 조건은 같은 "
                          "행, OR 조건은 다른 행에 적습니다. 열 순서는 달라도 "
                          "됩니다.")
        elif btype == "result":
            ok, msgs, order_diff = compare_result_block(ga, gs)
            if ok:
                passed_b += 1
                if order_diff:
                    res.notes.append("고급 필터 결과의 행 순서가 정답과 "
                                     "다르지만 내용이 같아 정답 처리했습니다.")
                continue
            res.details.append(f"[고급 필터 결과] {rng_txt}: "
                               + "; ".join(msgs))
            add_card(res, f"고급 필터 결과 ({rng_txt})",
                     res.alloc / len(blocks), "cell",
                     cells=[{"coord": rng_txt}],
                     props=[{"name": "결과 표", "expected": "지시 조건에 맞는 "
                             "행 추출", "got": m} for m in msgs[:3]],
                     hint="데이터 탭 → 고급에서 목록/조건/복사 위치 범위를 "
                          "다시 확인하세요.")
        else:
            cells3 = [(r, c, coord_of[(r, c)]) for (r, c) in cl]
            wrong = [(r, c, co) for (r, c, co) in cells3
                     if not judge.judge(r, c)[0]]
            if not wrong:
                passed_b += 1
                continue
            res.details.append(f"[결과 셀] {rng_txt}: {len(wrong)}개 불일치")
            add_wrong_cells(res, judge, wrong, limit=6)
            entries = make_cell_entries(judge, wrong, 3)
            add_card(res, f"결과 셀 ({rng_txt})", res.alloc / len(blocks),
                     "cell", cells=entries, more=max(0, len(wrong) - 3),
                     hint="텍스트 나누기/자동 필터 등 지시 결과를 원본과 "
                          "대조하세요.")
    res.earned = int(round(res.alloc * passed_b / len(blocks)))


CALC_FUNC_NOTE = ("이 채점기는 결과값 기준이라 문제가 제시한 함수를 썼는지는 "
                  "판정하지 못합니다. 실제 시험은 제시 함수 외 사용 시 값이 "
                  "맞아도 오답이니 문제지의 함수 지정을 확인하세요.")


def grade_calc(res, ctx):
    """계산작업: diff 셀 5그룹, 그룹 내 전 셀 일치 시에만 그룹 점수."""
    res.notes.append(CALC_FUNC_NOTE)
    diffs = ctx["vdiffs"]
    if not diffs:
        res.earned = res.alloc
        res.notes.append("문제/정답 간 차이가 없어 자동 만점 처리")
        return
    judge = ctx["judge"]
    key_groups = (ctx["key"].get("groups") or {}).get(norm_sheet_name(res.name))
    if key_groups:
        from openpyxl.utils import get_column_letter
        groups = []
        for g in key_groups:
            cells = []
            for ref in g:
                for (r, c) in expand_ref(ref):
                    cells.append((r, c, f"{get_column_letter(c)}{r}"))
            if cells:
                groups.append(cells)
    else:
        clusters = cluster_cells([(r, c) for (r, c, _) in diffs], CALC_GROUP_TARGET)
        co = {(r, c): coord for (r, c, coord) in diffs}
        groups = [[(r, c, co[(r, c)]) for (r, c) in cl] for cl in clusters]
    n = len(groups)
    per = res.alloc / n
    passed = 0
    for gi, cells in enumerate(groups, 1):
        wrong = [(r, c, co) for (r, c, co) in cells if not judge.judge(r, c)[0]]
        rng = f"{cells[0][2]}~{cells[-1][2]}" if len(cells) > 1 else cells[0][2]
        if not wrong:
            passed += 1
        else:
            res.details.append(
                f"[문제 {gi}] ({rng}) {len(cells)}개 셀 중 {len(wrong)}개 불일치 "
                f"→ 0/{per:.0f}점 (문제 단위 전부 정답 원칙)")
            add_wrong_cells(res, judge, wrong, limit=6)
            entries = make_cell_entries(judge, wrong, 3)
            rep_formula = next((e["formula"] for e in entries if e["formula"]),
                               None)
            if rep_formula is None:
                for (r, c, co) in cells:
                    ef = judge.judge(r, c)[3]
                    if ef:
                        rep_formula = ef
                        break
            add_card(res, f"계산 문제 {gi}", per, "cell", cells=entries,
                     formula=rep_formula, note=f"대상 범위 {rng}",
                     more=max(0, len(wrong) - 3),
                     hint="한 셀이라도 틀리면 문제 전체가 0점입니다. 첫 셀에 "
                          "수식을 완성한 뒤 채우기 핸들로 복사하고, 결과가 "
                          "이상한 셀이 없는지 훑어보세요.")
    res.earned = int(round(per * passed))
    res.notes.append(f"수식 문제를 {n}개 그룹으로 인식 (그룹당 {per:.0f}점)")


def grade_analysis(res, ctx):
    """분석작업: 확정 검사 항목(값/수식 diff) 전부 일치=만점, 아니면 0점.

    실제 시험 규정: 분석작업은 부분점수가 없습니다. 시나리오/피벗의
    '존재 여부' 같은 베스트 에포트 항목은 점수 게이트에서 제외하고
    노트로만 안내합니다 (자동 판정 한계로 인한 억울한 0점 방지).
    """
    book_a, book_s = ctx["book_a"], ctx["book_s"]
    asheet, ssheet = ctx["asheet"], ctx["ssheet"]
    diffs = ctx["vdiffs"]
    judge = ctx["judge"]
    if diffs:
        wrong = [(r, c, co) for (r, c, co) in diffs if not judge.judge(r, c)[0]]
        matched = len(diffs) - len(wrong)
        if wrong:
            res.earned = 0
            res.details.append(
                f"결과 셀 {len(diffs)}개 중 {matched}개 일치, {len(wrong)}개 "
                f"불일치 → 0점 (실제 시험 규정: 분석작업은 부분점수 없음)")
            add_wrong_cells(res, judge, wrong, limit=15)
            entries = make_cell_entries(judge, wrong, 3)
            add_card(res, "분석 결과", res.alloc, "cell", cells=entries,
                     formula=next((e["formula"] for e in entries
                                   if e["formula"]), None),
                     more=max(0, len(wrong) - 3),
                     hint="분석작업은 한 곳만 틀려도 0점입니다. 정렬 -> 부분합 "
                          "같은 지시 순서를 지키고, 완료 후 결과 표를 문제지의 "
                          "그림과 대조하세요.")
        else:
            res.earned = res.alloc
            res.details.append(f"결과 셀 {len(diffs)}개 전부 일치")
    else:
        res.earned = res.alloc
    # 시나리오 / 피벗: 존재 여부만 노트로 안내 (점수에 반영하지 않음)
    for has_a, has_s, label in (
        (book_a.has_scenarios(asheet),
         ssheet and book_s.has_scenarios(ssheet), "시나리오"),
        (book_a.has_pivot_for_sheet(asheet),
         ssheet and book_s.has_pivot_for_sheet(ssheet), "피벗 테이블"),
    ):
        if has_a:
            if not has_s:
                res.notes.append(
                    f"[{label}] 학생 파일에서 {label}을(를) 찾지 못했습니다. "
                    "자동 감점은 하지 않으니 직접 확인하세요.")
            res.notes.append(f"{label} 세부 설정은 자동 비교하지 않습니다. "
                             "수동 확인을 권장합니다.")


def grade_macro(res, ctx):
    """매크로작업: 결과 값+서식 diff 비례(8점 분량) + 단추/도형 존재 +2점."""
    book_a, book_s = ctx["book_a"], ctx["book_s"]
    asheet, ssheet = ctx["asheet"], ctx["ssheet"]
    judge = ctx["judge"]
    shs = book_s.raw[ssheet] if ssheet else None
    sha = book_a.raw[asheet]
    result_pts = max(0, res.alloc - 2)
    units_total = 0
    units_ok = 0
    # 값 diff
    wrong_cells = []
    for (r, c, co) in ctx["vdiffs"]:
        units_total += 1
        if judge.judge(r, c)[0]:
            units_ok += 1
        else:
            wrong_cells.append((r, c, co))
    # 서식 diff (셀 단위, 병합 인지 + edge 테두리)
    from openpyxl.utils.cell import coordinate_to_tuple
    a_members, a_anchors = merge_maps(sha)
    s_members, s_anchors = merge_maps(shs) if shs is not None else ({}, {})
    s_edges = None
    if "border" in ctx["fdiffs"] and shs is not None:
        _mr, _mc = scan_bounds(sha, shs)
        s_edges = sheet_edge_map(shs, _mr, _mc)
    fmt_bad = []
    for kind, payload in ctx["fdiffs"].items():
        if kind in ("rowheight", "colwidth", "merge", "names",
                    "border_edges"):
            continue
        if kind == "border":
            edge_info = ctx["fdiffs"].get("border_edges") or {}
            for coord in payload:
                units_total += 1
                ok = s_edges is not None and all(
                    (s_edges.get(key) or None) == (style or None)
                    for key, style in edge_info.get(coord, []))
                if ok:
                    units_ok += 1
                else:
                    fmt_bad.append(f"{coord}(테두리)")
            continue
        for coord in payload:
            units_total += 1
            try:
                r, c = coordinate_to_tuple(coord)
                ok = shs is not None and \
                    fmt_signature_m(sha, r, c, kind, a_members, a_anchors) \
                    == fmt_signature_m(shs, r, c, kind, s_members, s_anchors)
            except Exception:
                ok = False
            if ok:
                units_ok += 1
            else:
                fmt_bad.append(f"{coord}({FMT_KIND_LABEL.get(kind, kind)})")
    if units_total:
        earned = int(round(result_pts * units_ok / units_total))
    else:
        earned = result_pts
        res.notes.append("매크로 결과에 대한 diff가 없어 결과 점수 자동 부여")
    if wrong_cells:
        res.details.append(f"매크로 결과 셀 {len(wrong_cells)}개 불일치")
        add_wrong_cells(res, judge, wrong_cells, limit=10)
        entries = make_cell_entries(judge, wrong_cells, 3)
        add_card(res, "매크로 실행 결과",
                 max(0.0, result_pts - (earned if units_total else result_pts)),
                 "macro", cells=entries,
                 formula=next((e["formula"] for e in entries
                               if e["formula"]), None),
                 more=max(0, len(wrong_cells) - 3),
                 hint="매크로 기록(개발 도구 → 매크로 기록)을 시작한 뒤 지시 "
                      "작업을 수행하고 기록을 중지했는지, 기록 후 단추로 실행해 "
                      "결과가 나오는지 확인하세요.")
    if fmt_bad:
        by_kind = {}
        for b in fmt_bad:
            m = re.match(r"^([A-Z]+\d+)\((.+)\)$", b)
            if m:
                by_kind.setdefault(m.group(2), []).append(m.group(1))
        chips = [f"{k}: {', '.join(compress_coords(v)[:3])}"
                 for k, v in by_kind.items()] or fmt_bad[:5]
        res.details.append("매크로 서식 불일치: " + ", ".join(chips))
        add_card(res, "매크로 서식 결과", 0, "macro",
                 cells=[{"coord": ch, "expected": None, "got": None,
                         "formula": None} for ch in chips[:5]],
                 hint="서식 매크로가 지시한 서식(채우기·글꼴 등)을 정확히 "
                      "기록했는지 확인하세요.")
    # 단추/도형 존재 +2
    btn_a = book_a.sheet_has_drawing(asheet)
    btn_s = ssheet and book_s.sheet_has_drawing(ssheet)
    if btn_a:
        if btn_s:
            earned += 2
        else:
            res.details.append("[단추/도형] 매크로 실행용 단추(도형)를 찾지 못함 (-2점)")
            add_card(res, "매크로 단추(도형)", 2, "macro",
                     hint="개발 도구 → 삽입 → 단추(양식 컨트롤) 또는 삽입 → "
                          "도형을 그린 뒤 우클릭 → 매크로 지정으로 연결하고, "
                          "단추 텍스트를 지시대로 입력합니다.")
    else:
        earned += 2 if units_total == 0 or units_ok == units_total else 0
    res.earned = min(res.alloc, earned)
    # vba 보존 경고
    if book_a.has_vba() and not book_s.has_vba():
        res.notes.append("학생 파일에 매크로(vbaProject)가 없습니다. "
                         "매크로는 xlsm 형식으로 저장해야 보존됩니다.")
    res.notes.append("매크로 코드 자체가 아니라 실행 결과(값/서식)를 기준으로 "
                     "채점합니다.")


def grade_chart(res, ctx):
    """차트작업: 차트 XML 5개 항목 x (배점/5)."""
    book_a, book_s = ctx["book_a"], ctx["book_s"]
    asheet, ssheet = ctx["asheet"], ctx["ssheet"]
    fa = chart_features(book_a, asheet)
    fs = chart_features(book_s, ssheet) if ssheet else None
    per = res.alloc / 5.0
    if fa is None:
        res.earned = res.alloc
        res.notes.append("정답 파일에서 차트를 찾지 못해 자동 만점 처리")
        return
    if fs is None:
        res.earned = 0
        res.details.append("학생 파일에서 차트를 찾지 못했습니다 (0점)")
        add_card(res, "차트 없음", res.alloc, "chart",
                 hint="차트가 저장되지 않았습니다. 문제의 기본 차트를 지우지 "
                      "말고 그 위에서 지시 사항만 수정하세요.")
        return
    passed = 0

    def fail(label, exp_txt, got_txt, hint):
        res.details.append(f"[{label}] 기대 {exp_txt} / 학생 {got_txt}")
        add_card(res, f"차트 - {label}", per, "chart",
                 cells=[{"coord": None, "expected": str(exp_txt),
                         "got": str(got_txt), "formula": None}],
                 hint=hint)

    # 1. 존재 + 종류
    if fa["types"] == fs["types"]:
        passed += 1
    else:
        fail("차트 종류", fa["types"], fs["types"],
             "차트 영역 우클릭 → 차트 종류 변경. 콤보(혼합)형이면 계열별 "
             "차트 종류와 보조 축 체크를 확인하세요.")
    # 2. 제목
    ta = (fa["title"][1] if fa["title"] else None)
    ts = (fs["title"][1] if fs["title"] else None)
    if (ta is None) == (ts is None) and \
            (ta is None or ta.replace(" ", "") == (ts or "").replace(" ", "")):
        passed += 1
    else:
        fail("차트 제목", ta if ta is not None else "(제목 없음)",
             ts if ts is not None else "(제목 없음)",
             "차트 선택 → 차트 요소(+) → 차트 제목 체크 후 텍스트를 "
             "입력합니다(제거 지시면 체크 해제).")
    # 3. 계열 수 + 참조
    refs_a = sorted(x[0] or "" for x in fa["series"])
    refs_s = sorted(x[0] or "" for x in fs["series"])
    if len(fa["series"]) == len(fs["series"]) and refs_a == refs_s:
        passed += 1
    else:
        fail("데이터 계열", f"{len(fa['series'])}개 {refs_a}",
             f"{len(fs['series'])}개 {refs_s}",
             "차트 우클릭 → 데이터 선택에서 계열을 문제 지시대로 추가/제거 "
             "하고 각 계열의 값 범위를 확인합니다.")
    # 4. 데이터 레이블
    dl_a = sorted((x[0] or "") for x in fa["series"] if x[1])
    dl_s = sorted((x[0] or "") for x in fs["series"] if x[1])
    if dl_a == dl_s:
        passed += 1
    else:
        fail("데이터 레이블", f"계열 {dl_a or '없음'}", f"계열 {dl_s or '없음'}",
             "레이블을 붙일 계열만 한 번 클릭해 선택한 뒤 차트 요소(+) → "
             "데이터 레이블에서 위치를 지정합니다.")
    # 5. 축 설정(max/min/majorUnit) 있으면 축, 없으면 범례 위치
    ax_a = [x for x in fa["axes"] if any(v is not None for v in x)]
    if ax_a:
        ax_s = [x for x in fs["axes"] if any(v is not None for v in x)]
        if sorted(map(str, ax_a)) == sorted(map(str, ax_s)):
            passed += 1
        else:
            fail("축 설정", ax_a, ax_s,
                 "세로 축 더블클릭 → 축 서식에서 최소/최대 경계와 단위 "
                 "(주 단위)를 지시 값으로 입력합니다.")
    else:
        if fa["legend"] == fs["legend"]:
            passed += 1
        else:
            fail("범례 위치", fa["legend"], fs["legend"],
                 "차트 요소(+) → 범례에서 위치(아래쪽/오른쪽 등)를 "
                 "지정합니다.")
    res.earned = int(round(per * passed))
    if passed < 5:
        res.details.insert(0, f"차트 검사 5개 항목 중 {passed}개 통과 "
                              f"(항목당 {per:.0f}점)")


def grade_generic(res, ctx):
    """비표준 시트: 값 diff 비례."""
    diffs = ctx["vdiffs"]
    if not diffs:
        res.earned = res.alloc
        return
    judge = ctx["judge"]
    wrong = [(r, c, co) for (r, c, co) in diffs if not judge.judge(r, c)[0]]
    res.earned = int(round(res.alloc * (len(diffs) - len(wrong)) / len(diffs)))
    if wrong:
        res.details.append(f"diff 셀 {len(diffs)}개 중 {len(wrong)}개 불일치")
        add_wrong_cells(res, judge, wrong, limit=15)
        entries = make_cell_entries(judge, wrong, 3)
        add_card(res, "오답 셀", res.alloc - res.earned, "cell",
                 cells=entries, more=max(0, len(wrong) - 3),
                 formula=next((e["formula"] for e in entries
                               if e["formula"]), None))


GRADERS = {
    "기본작업-1": grade_basic1,
    "기본작업-2": grade_basic2,
    "기본작업-3": grade_basic3,
    "계산작업": grade_calc,
    "분석작업-1": grade_analysis,
    "분석작업-2": grade_analysis,
    "매크로작업": grade_macro,
    "차트작업": grade_chart,
}

NEEDS_FMT_DIFF = {"기본작업-2", "매크로작업"}


# ---------------------------------------------------------------------------
# 전체 채점
# ---------------------------------------------------------------------------


def run_grading(problem_path, answer_path, student_path, key, sheets=None):
    """채점 실행. sheets가 주어지면 해당 시트만 부분 채점.

    sheets: 시트명 목록(정규화 매칭). 없는 이름이면 RuntimeError.
    """
    book_p = Book(problem_path, "문제")
    book_a = Book(answer_path, "정답")
    book_s = Book(student_path, "학생")

    points = dict(DEFAULT_POINTS)
    for k, v in (key.get("points") or {}).items():
        points[norm_sheet_name(k)] = v

    global_notes = []
    results = []

    # 시트 매칭 (문제 기준)
    matched = []  # (norm_name, psheet, asheet, ssheet or None)
    for psheet in book_p.raw.sheetnames:
        n = norm_sheet_name(psheet)
        asheet = book_a.norm_map.get(n)
        if asheet is None:
            global_notes.append(f"정답 파일에 '{psheet}' 시트가 없어 채점에서 제외")
            continue
        ssheet = book_s.norm_map.get(n)
        matched.append((n, psheet, asheet, ssheet))
    for asheet in book_a.raw.sheetnames:
        n = norm_sheet_name(asheet)
        if n not in {m[0] for m in matched} and n not in \
                {norm_sheet_name(s) for s in book_p.raw.sheetnames}:
            global_notes.append(f"정답 전용 시트 '{asheet}'는 채점 대상이 아닙니다 "
                                "(예: 시나리오 요약)")

    if not matched:
        raise RuntimeError("문제/정답 파일에서 매칭되는 시트를 찾지 못했습니다.")

    # 부분 채점: 지정 시트만 남기기
    if sheets:
        wanted = [norm_sheet_name(s) for s in sheets if str(s).strip()]
        available = [m[0] for m in matched]
        unknown = [w for w in wanted if w not in available]
        if unknown:
            raise RuntimeError(
                "--sheets에 지정한 시트를 찾을 수 없습니다: "
                + ", ".join(unknown)
                + "\n사용 가능한 시트: " + ", ".join(available))
        matched = [m for m in matched if m[0] in wanted]
        global_notes.append("부분 채점: " + ", ".join(m[1] for m in matched)
                            + " 시트만 채점했습니다.")

    # diff 선계산 + 배점 할당
    ctxs = {}
    nonstd = []
    for (n, psheet, asheet, ssheet) in matched:
        shp, sha = book_p.raw[psheet], book_a.raw[asheet]
        vdiffs = value_diff_cells(shp, sha)
        fdiffs = format_diff_items(book_p, book_a, psheet, asheet) \
            if n in NEEDS_FMT_DIFF else {}
        key_cells = {k.upper(): v for k, v in
                     ((key.get("cells") or {}).get(n) or {}).items()}
        judge = CellJudge(book_a, book_s, asheet, ssheet, key_cells)
        ctxs[n] = {
            "book_p": book_p, "book_a": book_a, "book_s": book_s,
            "psheet": psheet, "asheet": asheet, "ssheet": ssheet,
            "vdiffs": vdiffs, "fdiffs": fdiffs, "judge": judge, "key": key,
        }
        if n not in points:
            nonstd.append(n)

    std_total = sum(points[n] for (n, *_ ) in matched if n in points)
    if nonstd:
        remain = max(0, 100 - std_total)
        weights = {n: max(1, len(ctxs[n]["vdiffs"])) for n in nonstd}
        wsum = sum(weights.values())
        base = remain if std_total else 100
        for n in nonstd:
            points[n] = base * weights[n] / wsum
        global_notes.append("표준 시트명이 아닌 시트는 diff 셀 수 비례로 "
                            "배점을 자동 배분했습니다.")

    formula_only = False
    for (n, psheet, asheet, ssheet) in matched:
        alloc = points.get(n, 0)
        res = SheetResult(psheet, alloc)
        ctx = ctxs[n]
        if ssheet is None:
            res.missing = True
            res.earned = 0
            res.details.append("학생 파일에 이 시트가 없습니다 (0점)")
            results.append(res)
            continue
        grader = GRADERS.get(n, grade_generic)
        try:
            grader(res, ctx)
        except Exception as e:
            res.notes.append(f"채점 중 오류가 발생해 값 비교로 대체: {e}")
            try:
                grade_generic(res, ctx)
            except Exception:
                pass
        res.earned = max(0, min(res.alloc if isinstance(res.alloc, int)
                                else res.alloc, res.earned))
        if ctx["judge"].formula_only_used:
            formula_only = True
        results.append(res)

    if formula_only:
        global_notes.append("일부 셀은 계산된 값이 없어 수식 기준으로 채점했습니다. "
                            "Excel에서 파일을 열어 저장하면 값 기준 채점이 가능합니다.")

    total_alloc = sum(r.alloc for r in results)
    total_earned = sum(r.earned for r in results)
    if sheets:
        # 부분 채점: 환산 없이 영역 배점 합 기준 원점수
        score100 = int(round(total_earned))
    elif total_alloc and abs(total_alloc - 100) > 0.01:
        score100 = round(total_earned / total_alloc * 100)
        global_notes.append(f"채점 대상 배점 합계가 {total_alloc:.0f}점이라 "
                            "100점 만점으로 환산했습니다.")
    else:
        score100 = int(round(total_earned))
    try:
        enrich_wrong_cards(results)   # 오답 카드에 유형 분류 + 수식 해설
    except Exception:
        pass
    return results, score100, global_notes, (book_p, book_a, book_s)


# ---------------------------------------------------------------------------
# 수식 해설기 (rule-based)
# ---------------------------------------------------------------------------

_CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮"

_FTOKEN = re.compile(r'''
    (?P<str>"(?:[^"]|"")*")
  | (?P<funcopen>[A-Za-z_][A-Za-z0-9_.]*\()
  | (?P<ref>\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?)
  | (?P<name>[A-Za-z가-힣_][A-Za-z가-힣0-9_.]*)
  | (?P<num>\d+(?:\.\d+)?%?)
  | (?P<op><>|<=|>=|[=<>+\-*/&^%])
  | (?P<comma>,)
  | (?P<open>\()
  | (?P<close>\))
  | (?P<ws>\s+)
''', re.VERBOSE)


def _parse_formula(f):
    """수식 문자열 -> 노드 트리. 노드: (kind, ...) 튜플."""
    s = str(f).lstrip("=").strip()
    s = re.sub(r"'[^']*'!", "", s)          # 시트 참조 제거
    s = re.sub(r"\[\d+\]", "", s)
    s = re.sub(r"[A-Za-z가-힣0-9_.]+!", "", s)
    toks = []
    for m in _FTOKEN.finditer(s):
        if m.lastgroup != "ws":
            toks.append((m.lastgroup, m.group()))
    pos = [0]

    def peek():
        return toks[pos[0]] if pos[0] < len(toks) else (None, None)

    def parse_expr():
        parts = []
        while True:
            k, v = peek()
            if k is None or k in ("comma", "close"):
                break
            if k == "funcopen":
                pos[0] += 1
                name = v[:-1].strip().upper().replace("_XLFN.", "")
                if name == "RANK":
                    name = "RANK.EQ"
                parts.append(("func", name, parse_args()))
            elif k == "open":
                pos[0] += 1
                inner = parse_expr()
                if peek()[0] == "close":
                    pos[0] += 1
                parts.append(inner)
            else:
                parts.append((k, v))
                pos[0] += 1
        if not parts:
            return ("empty", "")
        if len(parts) == 1:
            return parts[0]
        return ("expr", parts)

    def parse_args():
        args = []
        if peek()[0] == "close":
            pos[0] += 1
            return args
        while True:
            args.append(parse_expr())
            k, _v = peek()
            if k == "comma":
                pos[0] += 1
                continue
            if k == "close":
                pos[0] += 1
            break
        return args

    return parse_expr()


def _collect_funcs(node, out):
    if node[0] == "func":
        for a in node[2]:
            _collect_funcs(a, out)
        out.append(node)
    elif node[0] == "expr":
        for p in node[1]:
            _collect_funcs(p, out)


def _render(node, marks):
    kind = node[0]
    if kind == "func":
        if id(node) in marks:
            return marks[id(node)]
        return _fn_display(node, marks)
    if kind == "expr":
        return "".join(_render(p, marks) for p in node[1])
    if kind == "empty":
        return "(생략)"
    if kind == "ref":
        return node[1].replace("$", "")
    return node[1]


def _fn_display(node, marks):
    inner = ", ".join(_render(a, marks) for a in node[2])
    return f"{node[1]}({inner})"


def _lit_num(node):
    try:
        if node[0] == "num":
            v = node[1].rstrip("%")
            return float(v) if "." in v else int(v)
        if node[0] == "expr" and len(node[1]) == 2 and \
                node[1][0] == ("op", "-") and node[1][1][0] == "num":
            return -_lit_num(node[1][1])
    except Exception:
        pass
    return None


def _rng(text):
    return f"{text} 범위" if ":" in text else text


_DB_VERBS = {"DSUM": "합계", "DAVERAGE": "평균", "DCOUNT": "숫자 셀 개수",
             "DCOUNTA": "개수", "DMAX": "최대값", "DMIN": "최소값",
             "DGET": "값", "DPRODUCT": "곱"}
_AGG_KO = {"SUM": "합계", "AVERAGE": "평균", "MAX": "최대값", "MIN": "최소값",
           "MEDIAN": "중간값(중앙값)", "COUNT": "숫자 셀 개수",
           "COUNTA": "비어 있지 않은 셀 개수", "STDEV": "표준편차",
           "MODE": "최빈값", "MODE.SNGL": "최빈값", "VAR": "분산"}


def _round_pos(node):
    v = _lit_num(node)
    if v is None:
        return "지정한 자리로"
    v = int(v)
    if v > 0:
        return f"소수 {v}째 자리까지"
    if v == 0:
        return "일의 자리까지(정수로)"
    unit = {-1: "십", -2: "백", -3: "천", -4: "만"}.get(v, f"10^{-v}")
    return f"{unit} 단위로"


def _fn_sentence(name, args, R):
    n = len(args)

    def A(i, default=""):
        return R(args[i]) if i < n else default

    if name == "IF":
        return (f"조건 {A(0)} 이(가) 참이면 {A(1)}, 아니면 "
                f"{A(2, '빈 값')} 을(를) 표시합니다.")
    if name == "IFS":
        pairs = [f"{A(i)} → {A(i + 1)}" for i in range(0, n - 1, 2)]
        return ("조건을 앞에서부터 차례로 검사해 처음 참이 되는 조건의 값을 "
                "표시합니다: " + ", ".join(pairs))
    if name == "AND":
        return f"{', '.join(A(i) for i in range(n))} 조건이 모두 참인지 확인합니다."
    if name == "OR":
        return (f"{', '.join(A(i) for i in range(n))} 조건 중 하나라도 참인지 "
                "확인합니다.")
    if name == "IFERROR":
        return f"{A(0)} 이(가) 오류이면 {A(1)} 을(를) 대신 표시합니다."
    if name in ("VLOOKUP", "HLOOKUP"):
        where = "첫 열(세로)" if name == "VLOOKUP" else "첫 행(가로)"
        which = "열" if name == "VLOOKUP" else "행"
        mode = ""
        if n >= 4:
            last = A(3).upper()
            mode = (" (정확히 일치하는 값 찾기)" if last in ("0", "FALSE")
                    else " (구간/근사 일치: 찾는 값이 속하는 구간)")
        else:
            mode = " (구간/근사 일치: 찾는 값이 속하는 구간)"
        return (f"{A(0)} 값을 {_rng(A(1))}의 {where}에서 찾아, 같은 줄의 "
                f"{A(2)}번째 {which} 값을 가져옵니다{mode}.")
    if name == "INDEX":
        if n >= 3:
            return (f"{_rng(A(0))}에서 {A(1)}번째 행, {A(2)}번째 열이 만나는 "
                    "값을 가져옵니다.")
        return f"{_rng(A(0))}에서 {A(1)}번째 값을 가져옵니다."
    if name == "MATCH":
        mode = " (정확히 일치)" if n >= 3 and A(2) == "0" else ""
        return (f"{A(0)} 값이 {_rng(A(1))}에서 몇 번째에 있는지 위치(순번)를 "
                f"구합니다{mode}.")
    if name == "CHOOSE":
        opts = ", ".join(f"{i}이면 {A(i)}" for i in range(1, n))
        return f"{A(0)} 값이 {opts} 을(를) 선택합니다."
    if name == "LEFT":
        return f"{A(0)}의 왼쪽에서 {A(1, '1')}글자를 꺼냅니다."
    if name == "RIGHT":
        return f"{A(0)}의 오른쪽에서 {A(1, '1')}글자를 꺼냅니다."
    if name == "MID":
        return f"{A(0)}의 {A(1)}번째 글자부터 {A(2)}글자를 꺼냅니다."
    if name in ("SEARCH", "FIND"):
        extra = "(대소문자 구분 없음)" if name == "SEARCH" else "(대소문자 구분)"
        return f"{A(1)}에서 {A(0)} 이(가) 몇 번째 글자에 있는지 찾습니다{extra}."
    if name == "PROPER":
        return f"{A(0)}의 각 단어 첫 글자만 대문자로 바꿉니다."
    if name == "UPPER":
        return f"{A(0)}을(를) 모두 대문자로 바꿉니다."
    if name == "LOWER":
        return f"{A(0)}을(를) 모두 소문자로 바꿉니다."
    if name == "TRIM":
        return f"{A(0)}의 불필요한 공백을 제거합니다."
    if name == "LEN":
        return f"{A(0)}의 글자 수를 셉니다."
    if name in ("HOUR", "MINUTE", "SECOND"):
        ko = {"HOUR": "시", "MINUTE": "분", "SECOND": "초"}[name]
        return f"{A(0)}에서 '{ko}' 부분의 숫자만 꺼냅니다."
    if name in ("YEAR", "MONTH", "DAY"):
        ko = {"YEAR": "연도", "MONTH": "월", "DAY": "일"}[name]
        return f"{A(0)}에서 {ko} 숫자만 꺼냅니다."
    if name == "WEEKDAY":
        t = _lit_num(args[1]) if n >= 2 else None
        base = {2: " (옵션 2: 월요일=1 ~ 일요일=7)",
                3: " (옵션 3: 월요일=0 시작)"}.get(
            t, " (기본: 일요일=1 ~ 토요일=7)")
        return f"{A(0)}의 요일 번호를 구합니다{base}."
    if name == "MOD":
        return f"{A(0)}을(를) {A(1)}(으)로 나눈 나머지를 구합니다."
    if name == "QUOTIENT":
        return f"{A(0)}을(를) {A(1)}(으)로 나눈 몫(정수)을 구합니다."
    if name == "TODAY":
        return "오늘 날짜를 구합니다."
    if name == "NOW":
        return "현재 날짜와 시각을 구합니다."
    if name == "DATE":
        return f"연 {A(0)}, 월 {A(1)}, 일 {A(2)} 로 날짜를 만듭니다."
    if name == "TIME":
        return f"시 {A(0)}, 분 {A(1)}, 초 {A(2)} 로 시각을 만듭니다."
    if name in _DB_VERBS:
        return (f"{_rng(A(0))}(데이터 전체, 필드명 포함)에서 조건 범위 "
                f"{_rng(A(2))}의 조건에 맞는 행만 골라 {A(1)} 필드의 "
                f"{_DB_VERBS[name]}을(를) 구합니다.")
    if name == "COUNTIF":
        return f"{_rng(A(0))}에서 조건 {A(1)} 에 맞는 셀의 개수를 셉니다."
    if name == "COUNTIFS":
        pairs = [f"{_rng(A(i))}가 {A(i + 1)}" for i in range(0, n - 1, 2)]
        return "조건 " + " 그리고 ".join(pairs) + " 을(를) 모두 만족하는 행의 개수를 셉니다."
    if name == "SUMIF":
        if n >= 3:
            return (f"{_rng(A(0))}이 조건 {A(1)} 에 맞는 행의 {_rng(A(2))} "
                    "값을 모두 더합니다.")
        return f"{_rng(A(0))}에서 조건 {A(1)} 에 맞는 값을 모두 더합니다."
    if name in ("SUMIFS", "AVERAGEIFS"):
        verb = "더합니다" if name == "SUMIFS" else "평균을 구합니다"
        pairs = [f"{_rng(A(i))}가 {A(i + 1)}" for i in range(1, n - 1, 2)]
        return (f"조건 {' 그리고 '.join(pairs)} 을(를) 모두 만족하는 행의 "
                f"{_rng(A(0))} 값을 {verb}.")
    if name == "AVERAGEIF":
        if n >= 3:
            return (f"{_rng(A(0))}이 조건 {A(1)} 에 맞는 행의 {_rng(A(2))} "
                    "값의 평균을 구합니다.")
        return f"{_rng(A(0))}에서 조건 {A(1)} 에 맞는 값의 평균을 구합니다."
    if name == "LARGE":
        return f"{_rng(A(0))}에서 {A(1)}번째로 큰 값을 구합니다."
    if name == "SMALL":
        return f"{_rng(A(0))}에서 {A(1)}번째로 작은 값을 구합니다."
    if name == "RANK.EQ":
        order = ""
        if n >= 3 and _lit_num(args[2]) == 1:
            order = " (오름차순: 작을수록 1위)"
        else:
            order = " (내림차순: 클수록 1위)"
        return f"{A(0)} 값이 {_rng(A(1))}에서 몇 위인지 순위를 구합니다{order}."
    if name in _AGG_KO:
        rngs = ", ".join(_rng(A(i)) for i in range(n))
        return f"{rngs}의 {_AGG_KO[name]}을(를) 구합니다."
    if name == "ROUND":
        return f"{A(0)}을(를) {_round_pos(args[1] if n > 1 else None)} 반올림합니다."
    if name == "ROUNDUP":
        return f"{A(0)}을(를) {_round_pos(args[1] if n > 1 else None)} 올림합니다."
    if name == "ROUNDDOWN":
        return f"{A(0)}을(를) {_round_pos(args[1] if n > 1 else None)} 버림(내림)합니다."
    if name == "TRUNC":
        return f"{A(0)}의 소수 부분을 버립니다."
    if name == "INT":
        return f"{A(0)}보다 크지 않은 정수로 내립니다(소수 버림)."
    if name == "ABS":
        return f"{A(0)}의 절댓값(음수면 부호를 뗀 값)을 구합니다."
    if name == "POWER":
        return f"{A(0)}의 {A(1)}제곱을 구합니다."
    if name == "SUMPRODUCT":
        return "짝이 되는 값끼리 곱한 뒤 모두 더합니다."
    if name in ("CONCAT", "CONCATENATE"):
        return f"{', '.join(A(i) for i in range(n))} 을(를) 이어 붙입니다."
    if name == "TEXT":
        return f"{A(0)}을(를) 표시 형식 {A(1)} 문자열로 바꿉니다."
    args_txt = ", ".join(A(i) for i in range(n))
    return f"{name} 함수로 {args_txt} 을(를) 계산합니다."


def explain_formula(formula):
    """정답 수식을 안(내부)→밖 순서의 단계별 한국어 풀이로.

    반환: (단계 문자열 목록, 절대참조 포인트 문구 또는 None)
    """
    try:
        f = str(formula or "")
        if not f.startswith("="):
            return [], None
        if "__DATATABLE__" in f:
            return ["데이터 표 수식입니다. 데이터 탭 → 가상 분석 → 데이터 "
                    "표에서 행/열 입력 셀을 지정해 만듭니다."], None
        if "__ARRAY__" in f:
            return ["배열 수식입니다. 수식 입력 후 Ctrl+Shift+Enter로 "
                    "확정합니다."], None
        tree = _parse_formula(f)
        funcs = []
        _collect_funcs(tree, funcs)
        marks = {}
        steps = []
        for i, fn in enumerate(funcs[:15]):
            mark = _CIRCLED[i] if i < len(_CIRCLED) else f"({i + 1})"
            disp = _fn_display(fn, marks)
            sent = _fn_sentence(fn[1], fn[2], lambda nd: _render(nd, marks))
            steps.append(f"{mark} {disp}: {sent}")
            marks[id(fn)] = mark
        if tree[0] == "expr" and any(
                p[0] == "op" and p[1] == "&" for p in tree[1]):
            i = len(funcs)
            mark = _CIRCLED[i] if i < len(_CIRCLED) else f"({i + 1})"
            steps.append(f"{mark} {_render(tree, marks)}: 위 값들을 & 로 이어 "
                         "붙여 하나의 문자열로 만듭니다.")
        elif not funcs and tree[0] == "expr":
            steps.append(f"{_render(tree, marks)} 을(를) 계산합니다.")
        point = None
        if "$" in f:
            point = ("$ 절대참조는 채우기 핸들로 수식을 복사할 때 참조 범위가 "
                     "밀리지 않도록 고정하는 표시입니다. 표/기준 범위에만 $를 "
                     "붙이고, 행마다 바뀌어야 하는 셀에는 붙이지 않습니다.")
        return steps, point
    except Exception:
        return ["정답 수식을 한 단계씩 그대로 입력하며 구조를 익혀 보세요."], None


def explain_number_format(nf):
    """number_format 코드 문자열의 한국어 해설."""
    try:
        if not nf or nf == "General":
            return "기본 표시 형식(General)입니다."
        lits = [x for x in re.findall(r'"([^"]*)"', nf) if x.strip()]
        s = re.sub(r'"[^"]*"', "", nf)
        low = s.lower()
        feats = []
        if "#,##" in s or ",##0" in s:
            feats.append("천 단위 구분 기호(,) 표시")
        m = re.search(r"0\.(0+)", s)
        if m:
            feats.append(f"소수 {len(m.group(1))}자리까지 표시")
        if "%" in s:
            feats.append("백분율(%)로 표시")
        is_date = bool(re.search(r"[ymdhsa]", low))
        if not is_date:
            m = re.search(r"(?<![#,0.])0{2,}(?![#,0.])", s)
            if m:
                feats.append(f"숫자를 {len(m.group(0))}자리로 채워 표시"
                             "(빈 자리는 0)")
        if "@" in s:
            feats.append("입력한 텍스트는 그대로 두고")
        if is_date:
            bits = []
            if "yyyy" in low:
                bits.append("네 자리 연도")
            elif "yy" in low:
                bits.append("두 자리 연도")
            if re.search(r"(?<!a)m", low):
                bits.append("월(또는 분)")
            if re.search(r"(?<!d)d{1,2}(?!d)", low):
                bits.append("일")
            if "aaaa" in low:
                bits.append("한글 요일 전체(예: 월요일)")
            elif "aaa" in low:
                bits.append("한글 요일 약자(예: 월)")
            elif "ddd" in low:
                bits.append("영문 요일")
            if "h" in low:
                bits.append("시각")
            feats.append("날짜/시간을 " + "·".join(bits) + " 형태로 표시")
        for lit in lits:
            if is_date and lit.strip() in ("년", "월", "일", "시", "분", "초"):
                continue  # 날짜 설명에 이미 포함
            feats.append(f"문자 '{lit.strip()}' 붙여 표시")
        if not feats:
            return f"사용자 지정 형식 코드 {nf} 를 그대로 입력합니다."
        return f"형식 코드 {nf} 의 뜻: " + ", ".join(feats) + "."
    except Exception:
        return f"사용자 지정 형식 코드 {nf} 를 그대로 입력합니다."


# ---------------------------------------------------------------------------
# 수식 차이 진단기 (학생 수식 vs 정답 수식)
# ---------------------------------------------------------------------------


def _walk_funcs(node, depth=0, out=None):
    """트리에서 (함수명, 깊이, 노드) 목록 (DFS 선행 순서)."""
    if out is None:
        out = []
    if node[0] == "func":
        out.append((node[1], depth, node))
        for a in node[2]:
            _walk_funcs(a, depth + 1, out)
    elif node[0] == "expr":
        for p in node[1]:
            _walk_funcs(p, depth, out)
    return out


def _strip_strings(f):
    return re.sub(r'"[^"]*"', '""', str(f))


def _cmp_ops(f):
    """비교 연산자 목록 (문자열 리터럴 제외)."""
    return re.findall(r"<=|>=|<>|<|>", _strip_strings(f))


def _top_func_name(tree):
    if tree[0] == "func":
        return tree[1]
    if tree[0] == "expr":
        for p in tree[1]:
            if p[0] == "func":
                return p[1]
    return None


def _is_ref_node(n):
    return n[0] == "ref"


def _josa_wa(word):
    return f"{word}와(과)"


def diagnose_formula_diff(student_f, expected_f):
    """학생 수식과 정답 수식을 비교해 무엇이 다른지 한국어 진단 목록 생성.

    확실히 짚을 수 있는 차이만 나열하고, 없으면 '접근이 다릅니다' 안내.
    """
    fallback = ["정답 수식과 접근이 다릅니다. 아래 풀이를 따라 처음부터 "
                "다시 작성해 보세요."]
    try:
        sf, ef = str(student_f or ""), str(expected_f or "")
        if not sf.startswith("=") or not ef.startswith("="):
            return []

        def _abs_ref_notes():
            out = []
            e_anchored = set(re.findall(
                r"\$[A-Z]{1,3}\$\d+:\$[A-Z]{1,3}\$\d+", ef.upper()))
            for anch in sorted(e_anchored):
                plain = anch.replace("$", "")
                if re.search(re.escape(plain),
                             sf.upper().replace("$", "")) \
                        and anch not in sf.upper().replace(" ", ""):
                    out.append(f"{anch}처럼 범위를 $로 고정해야 채우기 "
                               "핸들로 복사할 때 범위가 어긋나지 않습니다.")
                    break
            return out

        if norm_formula(sf) == norm_formula(ef):
            # 정규화하면 같음 = $ 등 표기 차이만 -> 절대참조 진단만
            if re.sub(r"\s+", "", sf.upper()) != \
                    re.sub(r"\s+", "", ef.upper()):
                return _abs_ref_notes()
            return []
        notes = []
        st, et = _parse_formula(sf), _parse_formula(ef)
        sfuncs = _walk_funcs(st)
        efuncs = _walk_funcs(et)
        snames = [n for n, _d, _x in sfuncs]
        enames = [n for n, _d, _x in efuncs]
        sset, eset = set(snames), set(enames)
        s_top, e_top = _top_func_name(st), _top_func_name(et)
        # 1) 중첩 순서 반전 (예: =OR(IF(...)) vs =IF(OR(...)))
        if s_top and e_top and s_top != e_top \
                and s_top in eset and e_top in sset:
            notes.append(
                f"{_josa_wa(s_top)} {e_top}의 중첩 순서가 반대입니다 — "
                f"정답은 {e_top}(...) 안에 {s_top}(...)이 들어가는 구조입니다.")
        else:
            missing = sorted(eset - sset)
            extra = sorted(sset - eset)
            if missing:
                notes.append("정답에 있는 " + ", ".join(missing)
                             + " 함수가 내 수식에 없습니다.")
            if extra:
                notes.append("정답에 없는 " + ", ".join(extra)
                             + " 함수를 사용했습니다.")
        # 2) 비교 연산자 방향
        from collections import Counter
        sops, eops = Counter(_cmp_ops(sf)), Counter(_cmp_ops(ef))
        for a, b in ((">=", "<="), ("<=", ">="), (">", "<"), ("<", ">")):
            if sops[a] and eops[b] and not eops[a] and not sops[b]:
                notes.append(f"'{a}' 를 썼지만 '{b}' 여야 합니다 "
                             "(비교 방향이 반대).")
                break
        # 3) 절대참조 ($ 고정)
        notes.extend(_abs_ref_notes())
        # 4) 같은 함수의 인수 비교 (첫 등장 페어)
        seen = set()
        for name in enames:
            if name in seen or name not in sset:
                continue
            seen.add(name)
            sn = next(x for n, _d, x in sfuncs if n == name)
            en = next(x for n, _d, x in efuncs if n == name)
            sargs, eargs = sn[2], en[2]
            if len(sargs) != len(eargs):
                notes.append(f"{name}의 인수 개수가 다릅니다 "
                             f"(내 수식 {len(sargs)}개 / 정답 {len(eargs)}개).")
                continue
            for i, (sa, ea) in enumerate(zip(sargs, eargs), 1):
                rs, re_ = _render(sa, {}), _render(ea, {})
                if norm_formula("=" + rs) == norm_formula("=" + re_):
                    continue
                if sa[0] == "num" and ea[0] == "num":
                    notes.append(f"{name}의 {i}번째 인수가 {re_}이어야 "
                                 f"하는데 {rs}을(를) 썼습니다.")
                elif ea[0] == "str" and sa[0] in ("name", "ref") and \
                        ea[1].strip('"') == rs:
                    notes.append(f"문자는 따옴표로 감싸야 합니다: {ea[1]}")
                elif _is_ref_node(sa) and _is_ref_node(ea):
                    if rs.replace("$", "") != re_.replace("$", ""):
                        notes.append(f"{name}의 {i}번째 참조가 {re_}이어야 "
                                     f"하는데 {rs}을(를) 썼습니다.")
        # 5) & 연결 유무
        if "&" in _strip_strings(ef) and "&" not in _strip_strings(sf):
            notes.append("& 연결이 빠졌습니다 — 계산 결과 뒤에 &\"문자\" "
                         "형태로 이어 붙여야 합니다.")
        elif "&" in _strip_strings(sf) and "&" not in _strip_strings(ef):
            notes.append("정답에는 & 연결이 없습니다.")
        # 중복 제거 + 상한
        uniq = []
        for n in notes:
            if n not in uniq:
                uniq.append(n)
        return uniq[:6] if uniq else fallback
    except Exception:
        return fallback


# ---------------------------------------------------------------------------
# 취약점 유형 분류와 처방
# ---------------------------------------------------------------------------

CALC_CAT_FUNCS = [
    ("DB", {"DSUM", "DAVERAGE", "DCOUNT", "DCOUNTA", "DMAX", "DMIN", "DGET",
            "DPRODUCT", "DVAR", "DSTDEV"}),
    ("시간날짜", {"HOUR", "MINUTE", "SECOND", "YEAR", "MONTH", "DAY",
               "WEEKDAY", "DATE", "TIME", "TODAY", "NOW", "MOD", "EDATE",
               "EOMONTH", "DAYS", "DATEDIF"}),
    ("참조", {"VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "CHOOSE", "OFFSET",
            "LOOKUP", "XLOOKUP"}),
    ("문자열", {"LEFT", "MID", "RIGHT", "SEARCH", "FIND", "PROPER", "UPPER",
             "LOWER", "TRIM", "LEN", "REPT", "CONCAT", "CONCATENATE", "TEXT",
             "FIXED", "VALUE", "SUBSTITUTE", "REPLACE"}),
    ("판정", {"IF", "IFS", "AND", "OR", "NOT", "IFERROR", "SWITCH", "TRUE",
            "FALSE"}),
    ("통계조건", {"COUNTIF", "COUNTIFS", "SUMIF", "SUMIFS", "AVERAGEIF",
              "AVERAGEIFS", "LARGE", "SMALL", "RANK", "RANK.EQ", "MAX", "MIN",
              "MEDIAN", "COUNTA", "COUNT", "SUM", "AVERAGE", "MODE",
              "MODE.SNGL", "STDEV", "ROUND", "ROUNDUP", "ROUNDDOWN", "INT",
              "ABS", "TRUNC", "POWER", "SUMPRODUCT"}),
]

SHEET_CATEGORY = {
    "기본작업-1": "입력", "기본작업-2": "서식", "기본작업-3": "조건부서식·필터",
    "분석작업-1": "분석", "분석작업-2": "분석", "매크로작업": "매크로",
    "차트작업": "차트",
}

CAT_INFO = {
    "판정": {
        "이름": "판정(IF 계열)",
        "진단": "IF 중첩 순서와 AND/OR 조건식 구성에서 실수가 나옵니다.",
        "처방": ["계산드릴 '드릴1_판정' 시트 재풀이",
                 "루틴 페이지 실수 노트의 '판정' 항목에 이번 오답 수식 기록",
                 "IF·IFS·AND·OR 치트시트 복습 - 조건 검사 순서와 문자 결과의 "
                 "큰따옴표"],
    },
    "참조": {
        "이름": "찾기/참조(VLOOKUP 계열)",
        "진단": "찾을 범위 고정($)과 열/행 번호, 일치 옵션 지정이 약합니다.",
        "처방": ["계산드릴 '드릴2_참조' 시트 재풀이",
                 "루틴 페이지 실수 노트의 '참조' 항목에 이번 오답 수식 기록",
                 "VLOOKUP·HLOOKUP·INDEX/MATCH·CHOOSE 치트시트 복습 - 범위 "
                 "$고정과 정확히 일치(0) 옵션"],
    },
    "시간날짜": {
        "이름": "날짜/시간",
        "진단": "WEEKDAY 옵션, 시·분 추출, MOD 활용 등 날짜·시간 계산이 "
               "흔들립니다.",
        "처방": ["계산드릴 '드릴3_시간날짜' 시트 재풀이",
                 "루틴 페이지 실수 노트의 '시간날짜' 항목에 이번 오답 수식 기록",
                 "HOUR·MINUTE·YEAR·WEEKDAY·MOD 치트시트 복습 - WEEKDAY 두 번째 "
                 "인수(2=월요일 시작) 확인"],
    },
    "DB": {
        "이름": "데이터베이스 함수(D 계열)",
        "진단": "D함수의 세 인수(전체 범위·필드·조건 범위) 구조가 약합니다.",
        "처방": ["계산드릴 '드릴4_DB' 시트 재풀이",
                 "루틴 페이지 실수 노트의 'DB' 항목에 이번 오답 수식 기록",
                 "DSUM·DAVERAGE·DCOUNTA 치트시트 복습 - 데이터 범위는 필드명 "
                 "행부터, 조건 범위는 필드명+조건 세로 2칸"],
    },
    "통계조건": {
        "이름": "통계/조건 집계",
        "진단": "COUNTIF·SUMIF 계열의 조건 표기(따옴표·부등호)와 RANK·LARGE "
               "지정에서 실수가 나옵니다.",
        "처방": ["계산드릴 '드릴5_통계조건' 시트 재풀이",
                 "루틴 페이지 실수 노트의 '통계조건' 항목에 이번 오답 수식 기록",
                 "COUNTIF(S)·SUMIF(S)·RANK.EQ·LARGE/SMALL 치트시트 복습 - "
                 "조건은 \">=90\" 처럼 따옴표로"],
    },
    "문자열": {
        "이름": "문자열 처리",
        "진단": "글자 추출 위치·개수 지정과 & 연결 구성이 약합니다.",
        "처방": ["계산드릴 '드릴6_문자열' 시트 재풀이",
                 "루틴 페이지 실수 노트의 '문자열' 항목에 이번 오답 수식 기록",
                 "LEFT·MID·RIGHT·SEARCH·& 연결 치트시트 복습 - MID의 시작 "
                 "위치와 글자 수"],
    },
    "입력": {
        "이름": "자료 입력(기본작업-1)",
        "진단": "입력 오타·누락이 있습니다. 이 영역은 한 셀만 틀려도 5점 "
               "전체를 잃습니다.",
        "처방": ["입력을 마친 뒤 문제지와 셀 단위 1:1 대조(위→아래, 왼→오른쪽)",
                 "루틴 페이지 실수 노트에 틀린 셀과 원인(오타/누락) 기록"],
    },
    "서식": {
        "이름": "셀 서식(기본작업-2)",
        "진단": "표시 형식 코드·병합·테두리 등 서식 지시 이행이 누락됩니다.",
        "처방": ["셀 서식(Ctrl+1) 사용자 지정 형식 코드 5개 손으로 다시 쓰기 "
                 "(#,##0\"원\" 등)",
                 "루틴 페이지 실수 노트의 '서식' 항목에 틀린 항목 기록",
                 "기본작업-2 유형 1회분 재풀이"],
    },
    "조건부서식·필터": {
        "이름": "조건부 서식/고급 필터(기본작업-3)",
        "진단": "조건부 서식의 혼합참조($E5)나 고급 필터 조건 범위 작성이 "
               "약합니다.",
        "처방": ["조건부 서식 수식 규칙을 혼합참조로 3회 반복 작성",
                 "고급 필터 AND(같은 행)/OR(다른 행) 조건 배치 복습",
                 "루틴 페이지 실수 노트의 '기본작업-3' 항목에 기록"],
    },
    "분석": {
        "이름": "분석작업",
        "진단": "부분합·피벗·시나리오·목표값 등 분석 도구의 지시 순서 이행이 "
               "약합니다. 이 영역은 부분점수가 없습니다.",
        "처방": ["오답 난 분석 유형을 같은 파일에서 처음부터 다시 수행",
                 "정렬 → 부분합 순서 등 절차 암기(루틴 페이지 실수 노트에 기록)",
                 "완료 후 문제지의 결과 그림과 표를 1:1 대조"],
    },
    "매크로": {
        "이름": "매크로작업",
        "진단": "매크로 기록/단추 연결 절차가 미완성입니다.",
        "처방": ["매크로 기록 → 작업 → 기록 중지 → 단추에 연결 절차를 2회 반복",
                 "단추 텍스트·매크로 이름을 문제 지시와 똑같이 입력했는지 확인",
                 "xlsm 형식으로 저장하는 습관"],
    },
    "차트": {
        "이름": "차트작업",
        "진단": "차트 세부 요소(종류·제목·레이블·축) 설정이 누락됩니다.",
        "처방": ["차트 요소(+) 메뉴에서 제목·레이블·범례 넣고 빼기 연습",
                 "데이터 선택으로 계열 추가/제거 연습",
                 "루틴 페이지 실수 노트의 '차트' 항목에 틀린 요소 기록"],
    },
    "기타": {
        "이름": "기타",
        "진단": "그 밖의 항목에서 실점이 있습니다.",
        "처방": ["오답 항목을 같은 파일에서 다시 수행", "루틴 페이지 실수 노트에 기록"],
    },
}


def classify_formula_category(formula):
    """계산작업 정답 수식 -> 6유형 분류."""
    try:
        f = re.sub(r'"[^"]*"', "", str(formula or "").upper())
        names = set(re.findall(r"([A-Z][A-Z.]{1,15})\s*\(", f))
        names = {("RANK.EQ" if x == "RANK" else x).replace("_XLFN.", "")
                 for x in names}
        for cat, funcs in CALC_CAT_FUNCS:
            if cat == "DB" and names & funcs:
                return "DB"
        scores = {cat: len(names & funcs) for cat, funcs in CALC_CAT_FUNCS}
        if "&" in f:  # & 연결은 동점일 때만 문자열 쪽으로 기울도록 반 점
            scores["문자열"] = scores.get("문자열", 0) + 0.5
        best = max(scores.values() or [0])
        if best == 0:
            return "통계조건"
        for cat, _funcs in CALC_CAT_FUNCS:  # 동점이면 목록 순서 우선
            if scores.get(cat) == best:
                return cat
    except Exception:
        pass
    return "통계조건"


def enrich_wrong_cards(results):
    """카드에 유형 분류·수식 해설·수식 차이 진단을 채운다."""
    for r in results:
        nname = norm_sheet_name(r.name)
        for card in r.wrong:
            if nname == "계산작업":
                card["category"] = classify_formula_category(card.get("formula"))
            else:
                card["category"] = SHEET_CATEGORY.get(nname, "기타")
            card["expected_formula"] = card.get("formula")
            card["student_formula"] = next(
                (c.get("got_formula") for c in (card.get("cells") or [])
                 if c.get("got_formula")), None)
            card["diff_notes"] = []
            if card.get("formula"):
                steps, point = explain_formula(card["formula"])
                card["explain"] = steps
                if point:
                    card["point"] = point
                if card["student_formula"]:
                    card["diff_notes"] = diagnose_formula_diff(
                        card["student_formula"], card["formula"])


def build_diagnosis(results):
    """(유형별 실점 목록, TOP3 카드, 체크리스트) 생성."""
    losses = {}
    for r in results:
        for card in r.wrong:
            cat = card.get("category") or "기타"
            losses[cat] = losses.get(cat, 0) + (card.get("lost") or 0)
    ordered = sorted(losses.items(), key=lambda kv: -kv[1])
    top3 = []
    for cat, lost in ordered[:3]:
        info = CAT_INFO.get(cat, CAT_INFO["기타"])
        top3.append({"category": cat, "이름": info["이름"], "lost": lost,
                     "진단": info["진단"], "처방": info["처방"]})
    checklist = []
    has_calc_wrong = any(norm_sheet_name(r.name) == "계산작업" and r.wrong
                         for r in results)
    if has_calc_wrong:
        checklist.append("오답 난 계산 문제의 정답 수식을 보지 않고 손으로 "
                         "3번씩 다시 써 보기")
    for t in top3:
        if t["처방"]:
            checklist.append(t["처방"][0])
    checklist.append("내일 공부 첫 10분: 이 리포트의 오답노트만 다시 읽기")
    seen = set()
    uniq = []
    for c in checklist:
        if c not in seen:
            seen.add(c)
            uniq.append(c)
    return ordered, top3, uniq[:5]


# ---------------------------------------------------------------------------
# 출력
# ---------------------------------------------------------------------------


def print_console(results, score100, global_notes, paths, partial=False):
    line = "─" * 56
    print()
    print(line)
    print(" 컴활 2급 실기 자동 채점 결과")
    print(line)
    print(f" 문제 파일 : {os.path.basename(paths[0])}")
    print(f" 정답 파일 : {os.path.basename(paths[1])}")
    print(f" 학생 파일 : {os.path.basename(paths[2])}")
    print(line)
    print(" " + pad("시트", 22) + pad("배점", 8, "right")
          + pad("득점", 8, "right") + pad("판정", 10, "right"))
    print(line)
    for r in results:
        alloc = f"{r.alloc:.0f}"
        if r.missing:
            verdict = "시트없음"
        elif r.earned >= r.alloc:
            verdict = "정답"
        elif r.earned > 0:
            verdict = "부분정답"
        else:
            verdict = "오답"
        print(" " + pad(r.name, 22) + pad(alloc, 8, "right")
              + pad(f"{r.earned:.0f}", 8, "right") + pad(verdict, 10, "right"))
    print(line)
    max_alloc = sum(r.alloc for r in results)
    if partial:
        names = ", ".join(r.name for r in results)
        print(" " + pad("부분 채점 합계", 22)
              + pad(f"{max_alloc:.0f}", 8, "right")
              + pad(str(score100), 8, "right"))
        print(f" 부분 채점: {names} ({max_alloc:.0f}점 만점)")
    else:
        verdict = "합격권" if score100 >= PASS_LINE else "미달"
        print(" " + pad("총점", 22) + pad("100", 8, "right")
              + pad(str(score100), 8, "right"))
        print(f" 합격선 {PASS_LINE}점 기준: {verdict}")
    print(line)

    detailed = [r for r in results if r.details or r.notes]
    if detailed:
        print()
        print(" [상세 내역]")
        for r in detailed:
            print(f" ▷ {r.name} ({r.earned:.0f}/{r.alloc:.0f}점)")
            for d in r.details:
                print(f"    - {d}")
            for nt in r.notes:
                print(f"    * (참고) {nt}")
    if global_notes:
        print()
        print(" [안내]")
        for nt in global_notes:
            print(f"    * {nt}")
    print()


def esc(s):
    return html_mod.escape(str(s))


def _svg_donut(score, pass_line=PASS_LINE, max_total=100, show_pass=True):
    """총점 도넛 게이지. 부분 채점이면 영역 만점 기준, 합격선 눈금 생략."""
    r = 56
    c = 2 * math.pi * r
    max_total = max_total or 100
    frac = max(0.0, min(1.0, score / float(max_total)))
    if show_pass:
        color = "#107C41" if score >= pass_line else "#B45309"
    else:
        color = "#107C41"
    tick = ""
    if show_pass:
        ang = 2 * math.pi * (pass_line / 100.0) - math.pi / 2
        x1 = 70 + (r - 10) * math.cos(ang)
        y1 = 70 + (r - 10) * math.sin(ang)
        x2 = 70 + (r + 10) * math.cos(ang)
        y2 = 70 + (r + 10) * math.sin(ang)
        tick = (f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" '
                f'y2="{y2:.1f}" stroke="#B45309" stroke-width="2"/>')
    return (
        f'<svg width="150" height="150" viewBox="0 0 140 140" role="img" '
        f'aria-label="점수 {score}/{max_total}점">'
        f'<circle cx="70" cy="70" r="{r}" fill="none" stroke="#E8EFEA" '
        f'stroke-width="13"/>'
        f'<circle cx="70" cy="70" r="{r}" fill="none" stroke="{color}" '
        f'stroke-width="13" stroke-dasharray="{c * frac:.1f} {c:.1f}" '
        f'transform="rotate(-90 70 70)"/>' + tick +
        f'<text x="70" y="68" text-anchor="middle" font-size="32" '
        f'font-weight="700" fill="{color}" '
        f'style="font-variant-numeric:tabular-nums">{score}</text>'
        f'<text x="70" y="90" text-anchor="middle" font-size="12" '
        f'fill="#57705F">/ {max_total:.0f}점</text></svg>')


def _svg_sheet_bars(results):
    """시트별 획득/배점 가로 막대."""
    lab_w, bar_w, row_h, pad = 150, 380, 30, 10
    width = lab_w + bar_w + 90
    height = pad * 2 + row_h * len(results)
    rows = []
    for i, r in enumerate(results):
        y = pad + i * row_h
        cy = y + row_h / 2 + 4
        frac = (r.earned / r.alloc) if r.alloc else 0
        fill_w = bar_w * max(0.0, min(1.0, frac))
        zero = r.earned <= 0
        lab_color = "#B45309" if zero else "#1B3A26"
        bars = (f'<rect x="{lab_w}" y="{y + 7}" width="{bar_w}" '
                f'height="{row_h - 14}" rx="4" fill="#ECF1ED"/>')
        if fill_w > 0.5:
            bars += (f'<rect x="{lab_w}" y="{y + 7}" width="{fill_w:.1f}" '
                     f'height="{row_h - 14}" rx="4" fill="#107C41"/>')
        else:
            bars += (f'<rect x="{lab_w}" y="{y + 7}" width="5" '
                     f'height="{row_h - 14}" fill="#B45309"/>')
        rows.append(
            f'<text x="{lab_w - 8}" y="{cy}" text-anchor="end" font-size="12" '
            f'fill="{lab_color}">{esc(r.name)}</text>' + bars +
            f'<text x="{lab_w + bar_w + 8}" y="{cy}" font-size="12" '
            f'font-weight="600" fill="{lab_color}" '
            f'style="font-variant-numeric:tabular-nums">'
            f'{r.earned:.0f} / {r.alloc:.0f}</text>')
    return (f'<svg width="{width}" height="{height}" viewBox="0 0 {width} '
            f'{height}" role="img" aria-label="영역별 점수">'
            + "".join(rows) + "</svg>")


def _svg_trend(points):
    """과거 -> 이번 점수 꺾은선. points: [(라벨, 점수)]"""
    width, height = 560, 200
    ml, mr, mt, mb = 42, 18, 16, 34
    pw, ph = width - ml - mr, height - mt - mb
    n = len(points)

    def px(i):
        return ml + (pw * (i / (n - 1)) if n > 1 else pw / 2)

    def py(v):
        return mt + ph * (1 - max(0, min(100, v)) / 100.0)

    parts = []
    for g in (0, 25, 50, 75, 100):
        parts.append(f'<line x1="{ml}" y1="{py(g):.1f}" x2="{width - mr}" '
                     f'y2="{py(g):.1f}" stroke="#E8EFEA" stroke-width="1"/>')
        parts.append(f'<text x="{ml - 7}" y="{py(g) + 4:.1f}" '
                     f'text-anchor="end" font-size="10" fill="#8AA093">'
                     f'{g}</text>')
    parts.append(f'<line x1="{ml}" y1="{py(PASS_LINE):.1f}" '
                 f'x2="{width - mr}" y2="{py(PASS_LINE):.1f}" '
                 f'stroke="#B45309" stroke-width="1.5" '
                 f'stroke-dasharray="5 4"/>')
    parts.append(f'<text x="{ml + 5}" y="{py(PASS_LINE) - 5:.1f}" '
                 f'text-anchor="start" font-size="10" fill="#B45309">'
                 f'합격선 {PASS_LINE}</text>')
    coords = [(px(i), py(s)) for i, (_lab, s) in enumerate(points)]
    poly = " ".join(f"{x:.1f},{y:.1f}" for x, y in coords)
    parts.append(f'<polyline points="{poly}" fill="none" stroke="#107C41" '
                 f'stroke-width="2.5"/>')
    for i, ((lab, s), (x, y)) in enumerate(zip(points, coords)):
        last = i == n - 1
        ring = 'stroke="#FFFFFF" stroke-width="1.5"' if last else ""
        parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" '
                     f'r="{5 if last else 3.5}" fill="#107C41" {ring}/>')
        parts.append(f'<text x="{x:.1f}" y="{y - 9:.1f}" text-anchor="middle" '
                     f'font-size="11" font-weight="{700 if last else 400}" '
                     f'fill="#1B3A26" '
                     f'style="font-variant-numeric:tabular-nums">{s}</text>')
        parts.append(f'<text x="{x:.1f}" y="{height - 12}" '
                     f'text-anchor="middle" font-size="10" fill="#57705F">'
                     f'{esc(lab)}</text>')
    return (f'<svg width="{width}" height="{height}" viewBox="0 0 {width} '
            f'{height}" role="img" aria-label="성적 추이">'
            + "".join(parts) + "</svg>")


def _svg_cat_bars(cat_losses):
    """유형별 실점 가로 막대 (호박색)."""
    lab_w, bar_w, row_h, pad = 170, 320, 28, 8
    width = lab_w + bar_w + 80
    height = pad * 2 + row_h * len(cat_losses)
    peak = max((v for _c, v in cat_losses), default=1) or 1
    rows = []
    for i, (cat, lost) in enumerate(cat_losses):
        y = pad + i * row_h
        cy = y + row_h / 2 + 4
        w = bar_w * (lost / peak)
        name = CAT_INFO.get(cat, {}).get("이름", cat)
        rows.append(
            f'<text x="{lab_w - 8}" y="{cy}" text-anchor="end" font-size="12" '
            f'fill="#1B3A26">{esc(name)}</text>'
            f'<rect x="{lab_w}" y="{y + 6}" width="{bar_w}" '
            f'height="{row_h - 12}" rx="4" fill="#F4EFE6"/>'
            f'<rect x="{lab_w}" y="{y + 6}" width="{max(w, 3):.1f}" '
            f'height="{row_h - 12}" rx="4" fill="#B45309"/>'
            f'<text x="{lab_w + bar_w + 8}" y="{cy}" font-size="12" '
            f'font-weight="600" fill="#B45309" '
            f'style="font-variant-numeric:tabular-nums">-{lost:.0f}점</text>')
    return (f'<svg width="{width}" height="{height}" viewBox="0 0 {width} '
            f'{height}" role="img" aria-label="유형별 실점">'
            + "".join(rows) + "</svg>")


def _history_points(history, set_name, current_score):
    """기록에서 같은 세트의 점수 추이 추출 + 이번 점수."""
    pts = []
    for rec in history or []:
        if rec.get("mode") == "부분연습":  # 부분 연습 점수는 추이에서 제외
            continue
        nm = str(rec.get("세트명", ""))
        if set_name and (nm == set_name or nm.startswith(set_name)):
            sc = rec.get("점수")
            if isinstance(sc, (int, float)):
                stamp = str(rec.get("일시", ""))
                lab = stamp[5:10].replace("-", "/") if len(stamp) >= 10 else ""
                pts.append((lab or "이전", int(sc)))
    pts = pts[-8:]
    pts.append(("이번", int(current_score)))
    return pts


def _wrong_card_html(card):
    """오답노트 카드 1장."""
    cat = card.get("category")
    cat_chip = (f'<span class="chip">{esc(CAT_INFO.get(cat, {}).get("이름", cat))}'
                f'</span>') if cat else ""
    lost = card.get("lost") or 0
    lost_txt = f"-{lost:g}점" if lost else "감점 요인"
    head = (f'<div class="wc-top"><span class="wc-sheet">{esc(card["sheet"])}'
            f'</span><span class="wc-label">{esc(card["label"])}</span>'
            f'{cat_chip}<span class="wc-lost">{lost_txt}</span></div>')
    body = []
    cells = card.get("cells") or []
    chip_only = cells and all(c.get("expected") is None and c.get("got") is None
                              for c in cells)
    if chip_only:
        chips = "".join(f'<span class="cellchip">{esc(c["coord"])}</span>'
                        for c in cells)
        body.append(f'<div class="wc-chips">확인할 위치: {chips}'
                    + (f' 외 {card["more"]}곳' if card.get("more") else "")
                    + "</div>")
    else:
        for c in cells:
            got = c.get("got")
            gotf = c.get("got_formula")
            exp = c.get("expected")
            mine = esc(got if got is not None else "(비어 있음)")
            if gotf and str(gotf) != str(got):
                mine += f'<span class="sub">{esc(gotf)}</span>'
            ans = esc(exp if exp is not None else "(비어 있음)")
            coord = f'<span class="coord">{esc(c["coord"])}</span>' \
                if c.get("coord") else ""
            body.append(
                f'<div class="cmp">{coord}'
                f'<div class="box mine"><span class="tag">내 답</span>'
                f'<code>{mine}</code></div>'
                f'<div class="arrow">→</div>'
                f'<div class="box ans"><span class="tag">정답</span>'
                f'<code>{ans}</code></div></div>')
        if card.get("more"):
            body.append(f'<div class="wc-chips">외 {card["more"]}개 셀 동일 유형'
                        "</div>")
    if card.get("props"):
        rows = "".join(
            f'<div class="prop-row"><span class="prop-name">'
            f'{esc(p.get("name", "?"))}</span>'
            f'<span class="prop-exp">정답: {esc(p.get("expected", "?"))}</span>'
            f'<span class="prop-got">내 답: {esc(p.get("got", "?"))}</span>'
            "</div>"
            for p in card["props"][:8])
        body.append(f'<div class="props">{rows}</div>')
    if card.get("note"):
        body.append(f'<p class="wc-note">{esc(card["note"])}</p>')
    if card.get("diff_notes"):
        diffs = "".join(f'<div class="diff-item">{esc(d)}</div>'
                        for d in card["diff_notes"])
        body.append('<div class="diffbox"><div class="diff-title">무엇이 '
                    "다른가</div>" + diffs + "</div>")
    solve = []
    if card.get("formula"):
        solve.append(f'<code class="formula">{esc(card["formula"])}</code>')
    for step in card.get("explain") or []:
        solve.append(f'<div class="step">{esc(step)}</div>')
    if card.get("point"):
        solve.append(f'<p class="point">포인트: {esc(card["point"])}</p>')
    if solve:
        body.append('<div class="solve"><div class="solve-title">정확한 풀이'
                    "</div>" + "".join(solve) + "</div>")
    if card.get("hint"):
        body.append(f'<p class="hint">방법: {esc(card["hint"])}</p>')
    return f'<div class="wc">{head}{"".join(body)}</div>'


def write_html(path, results, score100, global_notes, paths,
               set_name=None, history=None, partial=False):
    """학습 리포트 HTML 생성 (자체 완결, 인라인 SVG).

    partial=True면 선택 시트만 채점한 부분 리포트: 영역 배점 만점 기준
    도넛, 합격선/배지/성적 추이 생략.
    """
    verdict = "합격권" if score100 >= PASS_LINE else "미달"
    vcls = "pass" if score100 >= PASS_LINE else "fail"
    set_name = set_name or derive_set_name(paths[0])
    now_txt = datetime.now().strftime("%Y-%m-%d %H:%M")
    all_cards = [c for r in results for c in r.wrong]
    cat_losses, top3, checklist = build_diagnosis(results)
    max_total = sum(r.alloc for r in results) if partial else 100
    graded_names = ", ".join(r.name for r in results)

    # 성적 추이 (부분 채점은 표시하지 않음 — 시험 점수와 섞이지 않게)
    trend_html = ""
    if history and not partial:
        pts = _history_points(history, set_name, score100)
        if len(pts) >= 2:
            trend_html = ('<div class="card"><h2>성적 추이</h2>'
                          '<div class="scroll">' + _svg_trend(pts)
                          + "</div></div>")

    # 오답노트
    if all_cards:
        wrong_html = ('<div class="card"><h2>오답노트</h2>'
                      '<p class="lead">틀린 항목마다 내 답과 정답을 비교하고, '
                      '정답 수식의 풀이를 단계별로 설명합니다.</p>'
                      + "".join(_wrong_card_html(c) for c in all_cards)
                      + "</div>")
    else:
        wrong_html = ('<div class="card"><h2>오답노트</h2>'
                      '<p class="perfect">틀린 항목이 없습니다. 모든 채점 '
                      '항목을 통과했습니다.</p></div>')

    # 취약점 진단
    if cat_losses:
        tops = []
        for i, t in enumerate(top3, 1):
            items = "".join(f"<li>{esc(p)}</li>" for p in t["처방"])
            tops.append(
                f'<div class="weak"><div class="rank">{i}</div>'
                f'<div class="weak-body"><h4>{esc(t["이름"])} '
                f'<span class="weak-lost">-{t["lost"]:g}점</span></h4>'
                f'<p>{esc(t["진단"])}</p><ul>{items}</ul></div></div>')
        diag_html = ('<div class="card"><h2>취약점 진단</h2>'
                     '<div class="scroll">' + _svg_cat_bars(cat_losses)
                     + '</div><h3 class="sub-h">취약 TOP '
                     + str(len(top3)) + "</h3>" + "".join(tops) + "</div>")
    else:
        diag_html = ""

    # 다음 액션
    if not checklist:
        checklist = ["내일 다른 회차 모의고사 1세트로 컨디션 유지",
                     "헷갈렸던 함수 1개를 골라 치트시트 훑어보기"]
    check_html = ('<div class="card"><h2>다음 액션</h2><ul class="checklist">'
                  + "".join(f'<li><label><input type="checkbox"> {esc(c)}'
                            "</label></li>" for c in checklist)
                  + "</ul></div>")

    gnotes = "".join(f"<li>{esc(n)}</li>" for n in global_notes)
    sheet_notes = []
    for r in results:
        for nt in r.notes:
            sheet_notes.append(f"<li>[{esc(r.name)}] {esc(nt)}</li>")
    notes_html = ""
    if gnotes or sheet_notes:
        notes_html = ('<div class="card"><h2>채점 참고</h2><ul class="notes">'
                      + "".join(sheet_notes) + gnotes + "</ul></div>")

    css = """
* { box-sizing:border-box; margin:0; padding:0; }
body { background:#F4FAF5; color:#1B3A26;
  font-family:'Malgun Gothic','맑은 고딕',Pretendard,sans-serif;
  line-height:1.65; padding:28px 14px; }
.wrap { max-width:820px; margin:0 auto; }
.card { background:#FFFFFF; border:1px solid #D7E3DA; border-radius:14px;
  padding:26px 30px; margin-bottom:18px; }
h1 { font-size:1.3rem; color:#107C41; }
h2 { font-size:1.06rem; color:#107C41; margin-bottom:12px;
  padding-bottom:8px; border-bottom:1px solid #E8EFEA; }
.sub-h { font-size:0.98rem; margin:16px 0 10px; color:#1B3A26; }
.meta { color:#57705F; font-size:0.84rem; margin-top:2px; }
.head-flex { display:flex; gap:26px; align-items:center; flex-wrap:wrap; }
.head-info { flex:1; min-width:230px; }
.badge { display:inline-block; margin-top:10px; padding:5px 16px;
  border-radius:999px; font-weight:700; font-size:0.92rem; color:#fff; }
.badge.pass { background:#107C41; }
.badge.fail { background:#B3372E; }
.files { color:#57705F; font-size:0.8rem; margin-top:12px;
  word-break:break-all; }
.scroll { overflow-x:auto; }
svg { font-family:'Malgun Gothic','맑은 고딕',sans-serif; display:block; }
.lead { color:#57705F; font-size:0.86rem; margin-bottom:14px; }
.perfect { color:#107C41; font-weight:700; }
.wc { border:1px solid #D7E3DA; border-left:4px solid #B45309;
  border-radius:10px; padding:14px 16px; margin-bottom:14px; }
.wc-top { display:flex; align-items:center; gap:8px; flex-wrap:wrap;
  margin-bottom:10px; }
.wc-sheet { color:#57705F; font-size:0.8rem; }
.wc-label { font-weight:700; font-size:0.95rem; }
.chip { background:#E3F2E8; color:#0B5D31; font-size:0.74rem;
  padding:2px 10px; border-radius:999px; }
.wc-lost { margin-left:auto; color:#B45309; font-weight:700;
  font-variant-numeric:tabular-nums; }
.cmp { display:flex; align-items:stretch; gap:8px; margin-bottom:8px;
  flex-wrap:wrap; }
.coord { align-self:center; min-width:44px; font-weight:700;
  font-size:0.85rem; color:#1B3A26; font-variant-numeric:tabular-nums; }
.box { flex:1; min-width:180px; border-radius:8px; padding:7px 10px;
  border:1px solid #D7E3DA; }
.box.mine { background:#FBF6EE; border-color:#E5D5B8; }
.box.ans { background:#F0F7F2; border-color:#BFDCCB; }
.box .tag { display:block; font-size:0.7rem; color:#57705F;
  margin-bottom:2px; }
.box code { font-family:Consolas,'Courier New',monospace; font-size:0.84rem;
  word-break:break-all; }
.box .sub { display:block; font-size:0.76rem; color:#8A7B5E;
  font-family:Consolas,monospace; word-break:break-all; }
.arrow { align-self:center; color:#57705F; }
.wc-chips { font-size:0.84rem; color:#57705F; margin-bottom:8px; }
.cellchip { display:inline-block; background:#ECF1ED; border-radius:6px;
  padding:1px 8px; margin:0 3px 3px 0; font-family:Consolas,monospace;
  font-size:0.8rem; }
.wc-note { font-size:0.85rem; color:#1B3A26; background:#F4FAF5;
  border-radius:8px; padding:8px 12px; margin-bottom:8px; }
.props { border:1px solid #D7E3DA; border-radius:8px; padding:8px 12px;
  margin-bottom:8px; }
.prop-row { display:flex; gap:12px; flex-wrap:wrap; font-size:0.86rem;
  padding:3px 0; border-bottom:1px dashed #E8EFEA; }
.prop-row:last-child { border-bottom:none; }
.prop-name { font-weight:700; min-width:88px; }
.prop-exp { color:#0B5D31; }
.prop-got { color:#B45309; }
.diffbox { background:#FBF6EE; border:1px solid #E5D5B8; border-radius:8px;
  padding:10px 14px; margin-bottom:8px; }
.diff-title { font-weight:700; font-size:0.84rem; color:#8A5A00;
  margin-bottom:6px; }
.diff-item { font-size:0.86rem; margin-bottom:4px; padding-left:4px; }
.solve { background:#F7FBF8; border:1px solid #E1EDE5; border-radius:8px;
  padding:10px 14px; margin-bottom:8px; }
.solve-title { font-weight:700; font-size:0.84rem; color:#0B5D31;
  margin-bottom:6px; }
.formula { display:block; font-family:Consolas,'Courier New',monospace;
  font-size:0.85rem; background:#FFFFFF; border:1px solid #D7E3DA;
  border-radius:6px; padding:6px 10px; margin-bottom:8px;
  word-break:break-all; }
.step { font-size:0.86rem; margin-bottom:4px; padding-left:4px; }
.point { font-size:0.82rem; color:#0B5D31; background:#E3F2E8;
  border-radius:6px; padding:6px 10px; margin-top:8px; }
.hint { font-size:0.82rem; color:#57705F; }
.weak { display:flex; gap:14px; border:1px solid #D7E3DA; border-radius:10px;
  padding:14px 16px; margin-bottom:10px; }
.rank { width:30px; height:30px; border-radius:50%; background:#B45309;
  color:#fff; font-weight:700; display:flex; align-items:center;
  justify-content:center; flex:none; font-variant-numeric:tabular-nums; }
.weak-body h4 { font-size:0.95rem; margin-bottom:4px; }
.weak-lost { color:#B45309; font-size:0.84rem;
  font-variant-numeric:tabular-nums; }
.weak-body p { font-size:0.86rem; color:#1B3A26; margin-bottom:6px; }
.weak-body ul { padding-left:18px; font-size:0.84rem; color:#57705F; }
.weak-body li { margin-bottom:2px; }
.checklist { list-style:none; }
.checklist li { margin-bottom:8px; font-size:0.9rem; }
.checklist input { accent-color:#107C41; margin-right:8px;
  transform:scale(1.15); }
.notes { padding-left:18px; font-size:0.82rem; color:#57705F; }
.notes li { margin-bottom:3px; }
footer { color:#57705F; font-size:0.78rem; text-align:center;
  margin-top:6px; }
"""
    doc = (
        "<!doctype html>\n"
        '<html lang="ko"><head><meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        f"<title>학습 리포트 - {esc(set_name)}</title>\n"
        f"<style>{css}</style></head><body><div class=\"wrap\">\n"
        '<div class="card"><div class="head-flex">'
        f'<div>{_svg_donut(score100, max_total=max_total, show_pass=not partial)}</div>'
        '<div class="head-info">'
        + (f"<h1>{esc(set_name)} 부분 연습 리포트</h1>"
           f'<div class="meta">부분 채점: {esc(graded_names)} '
           f'({max_total:.0f}점 만점) · 채점 일시 {now_txt}</div>'
           if partial else
           f"<h1>{esc(set_name)} 학습 리포트</h1>"
           f'<div class="meta">채점 일시 {now_txt} · 합격선 {PASS_LINE}점</div>'
           f'<span class="badge {vcls}">{verdict}</span>')
        + '<div class="files">'
        f"문제 {esc(os.path.basename(paths[0]))}<br>"
        f"정답 {esc(os.path.basename(paths[1]))}<br>"
        f"풀이 {esc(os.path.basename(paths[2]))}</div>"
        "</div></div></div>\n"
        '<div class="card"><h2>영역별 점수</h2><div class="scroll">'
        + _svg_sheet_bars(results) + "</div></div>\n"
        + trend_html + wrong_html + diag_html + check_html + notes_html
        + f"<footer>코코 채점 학습 리포트 · diff 기반 자동 채점 · {now_txt}"
          "</footer>\n</div></body></html>")
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


def derive_set_name(problem_path):
    """문제 파일명에서 세트명 추출 (…_문제.xlsx -> …)."""
    base = os.path.splitext(os.path.basename(str(problem_path)))[0]
    for suf in ("_문제", "-문제", " 문제"):
        if base.endswith(suf):
            return base[: -len(suf)]
    return base


def write_json(path, results, score100, global_notes, paths, partial=False):
    data = {
        "total": score100,
        "pass_line": PASS_LINE,
        "passed": None if partial else score100 >= PASS_LINE,
        "mode": "partial" if partial else "full",
        "graded_sheets": [r.name for r in results],
        "max_total": round(sum(r.alloc for r in results), 1)
        if partial else 100,
        "generated": datetime.now().isoformat(timespec="seconds"),
        "files": {
            "problem": os.path.abspath(paths[0]),
            "answer": os.path.abspath(paths[1]),
            "student": os.path.abspath(paths[2]),
        },
        "sheets": [
            {
                "name": r.name,
                "alloc": round(r.alloc, 2),
                "earned": round(r.earned, 2),
                "missing": r.missing,
                "details": r.details,
                "notes": r.notes,
            }
            for r in results
        ],
        "notes": global_notes,
        "wrong_items": [
            {
                "sheet": card["sheet"], "label": card["label"],
                "lost": card["lost"], "kind": card["kind"],
                "category": card.get("category"),
                "cells": card.get("cells") or [],
                "formula": card.get("formula"),
                "note": card.get("note"), "hint": card.get("hint"),
                "explain": card.get("explain") or [],
                "point": card.get("point"),
                "student_formula": card.get("student_formula"),
                "expected_formula": card.get("expected_formula"),
                "diff_notes": card.get("diff_notes") or [],
                "props": card.get("props") or [],
            }
            for r in results for card in r.wrong
        ],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main(argv=None):
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass
    ap = argparse.ArgumentParser(
        description="컴활 2급 실기 자동 채점 (문제↔정답 diff 기반)")
    ap.add_argument("--problem", required=True, help="문제 파일(.xlsx/.xlsm)")
    ap.add_argument("--answer", required=True, help="정답 파일(.xlsx/.xlsm)")
    ap.add_argument("--student", required=True, help="학생 풀이 파일(.xlsx/.xlsm)")
    ap.add_argument("--key", help="기대값/배점/그룹 JSON (선택)")
    ap.add_argument("--html", help="HTML 리포트 저장 경로 (선택)")
    ap.add_argument("--json", dest="json_out",
                    help="채점 결과(총점·시트별 점수) JSON 저장 경로 (선택)")
    ap.add_argument("--history",
                    help="응시 기록 JSON(기록.json) 경로 - 있으면 HTML "
                         "리포트에 같은 세트의 성적 추이를 표시 (선택)")
    ap.add_argument("--sheets",
                    help="부분 채점: 지정 시트만 채점 (쉼표 구분, 예: "
                         "--sheets 계산작업 또는 --sheets 기본작업-1,"
                         "기본작업-2). 총점은 해당 시트 배점 합 기준 (선택)")
    args = ap.parse_args(argv)

    key = {}
    if args.key:
        try:
            with open(args.key, encoding="utf-8") as f:
                key = json.load(f)
        except Exception as e:
            print(f"경고: --key 파일을 읽을 수 없어 무시합니다: {e}")
    key = normalize_key(key)

    sheets = None
    if args.sheets:
        sheets = [s.strip() for s in str(args.sheets).split(",") if s.strip()]
    try:
        results, score100, notes, _books = run_grading(
            args.problem, args.answer, args.student, key, sheets=sheets)
    except (FileNotFoundError, RuntimeError) as e:
        print(f"오류: {e}")
        return 2

    partial = bool(sheets)
    paths = (args.problem, args.answer, args.student)
    print_console(results, score100, notes, paths, partial=partial)
    # 기본 산출물 경로: 명시하지 않으면 풀이 파일 옆 '채점결과/' 폴더
    set_name = derive_set_name(args.problem)
    html_path, json_path = args.html, args.json_out
    if not html_path or not json_path:
        out_dir = os.path.join(
            os.path.dirname(os.path.abspath(args.student)), "채점결과")
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        try:
            os.makedirs(out_dir, exist_ok=True)
            base = os.path.join(out_dir, f"채점결과_{set_name}_{stamp}")
            if not html_path:
                html_path = base + ".html"
            if not json_path:
                json_path = base + ".json"
        except OSError as e:
            print(f"경고: 채점결과 폴더를 만들 수 없어 리포트 저장을 "
                  f"생략합니다: {e}")
    history = None
    if args.history:
        try:
            with open(args.history, encoding="utf-8") as f:
                loaded = json.load(f)
            history = loaded if isinstance(loaded, list) else None
        except Exception as e:
            print(f"경고: --history 파일을 읽을 수 없어 무시합니다: {e}")
    if html_path:
        write_html(html_path, results, score100, notes, paths,
                   set_name=set_name, history=history, partial=partial)
        print(f"HTML 리포트를 저장했습니다: {html_path}")
    if json_path:
        write_json(json_path, results, score100, notes, paths,
                   partial=partial)
        print(f"JSON 결과를 저장했습니다: {json_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
